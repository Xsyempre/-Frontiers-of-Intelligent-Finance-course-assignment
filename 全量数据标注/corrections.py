# -*- coding: utf-8 -*-
"""
全量标注修正脚本
修正v2自动标注中的错误
"""

import sys
import io
import json
import re
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ============================================================
# 修正规则
# ============================================================

# ann_related 修正：将0改为1（漏判的年报相关行）
# 格式: row_idx -> (ann_related, ann_year, ann_fin_flag, ann_fin_info, third_party_flag, third_party_list)
ann_related_corrections = {
    # 未在法定期限内披露年度报告
    83: (1, [2018], 0, None, None, None),  # 华泽钴镍未披露2018年年报
    245: (1, None, 0, None, None, None),  # *ST新潮未按期披露定期报告
    355: (1, [2005], 0, None, None, None),  # 未按规定披露2005年年报

    # 年报中未披露关联交易/担保
    302: (1, [2012], 0, None, None, None),  # 大元股份未在2012年年报中如实披露关联交易

    # 年报中关联交易超出预计
    501: (1, None, 0, None, None, None),  # 年报中披露关联交易超出预计

    # 年报中未披露关联方
    3000: None,  # 需要检查

    # 业绩预告与年报差异
    246: (1, [2015], 1, [{"year": 2015, "elements": ["利润"]}], 0, None),  # 业绩预告与年报差异

    # 年报信息披露不及时
    285: (1, [2016], 0, None, None, None),  # 政府补助迟至年报才披露
}

# third_party 修正：将1改为0（误判的第三方）
# 这些行的"第三方"实际上是审计机构/评估机构的失职，不是配合造假
third_party_remove = set()

# 需要检查的行（审计机构作为处罚对象，不是配合造假的第三方）
audit_firm_rows = []


def find_audit_firm_rows(ws):
    """找出审计机构/评估机构作为处罚对象的行"""
    rows = []
    for r in range(2, ws.max_row + 1):
        tp = ws.cell(row=r, column=6).value
        act = str(ws.cell(row=r, column=1).value)
        if tp == 1:
            # 检查是否是审计机构/评估机构作为处罚对象
            # 模式：文本开头提到会计师事务所/评估公司
            if re.search(r'^(?:经查明,)?[一-龥]{2,20}会计师事务所', act[:80]):
                rows.append(r)
            elif re.search(r'^(?:经查明,)?[一-龥]{2,20}评估(?:有限)?公司', act[:80]):
                rows.append(r)
            elif re.search(r'^[一-龥]{2,20}资产评估', act[:80]):
                rows.append(r)
    return rows


def clean_third_party_list(tl_str):
    """清理第三方列表中的错误名称"""
    if not tl_str or tl_str == 'null':
        return tl_str

    try:
        parsed = json.loads(tl_str)
        if not parsed:
            return tl_str

        cleaned = []
        for item in parsed:
            name = item.get('name', '')

            # 跳过"未明确"
            if name == '未明确':
                continue

            # 清理名称中多余的前缀
            # 模式: "虚构与XXX公司" -> "XXX公司"
            m = re.search(r'(?:虚构|通过|与|和)([一-龥A-Za-z]{4,40}(?:有限公司|股份有限公司|合伙企业|集团)(?:\([一-龥]*\))?)', name)
            if m:
                name = m.group(1)

            # 模式: "实际控制XXX公司" -> "XXX公司"
            m = re.search(r'(?:实际控制|控股股东)([一-龥A-Za-z]{4,40}(?:有限公司|股份有限公司|合伙企业))', name)
            if m:
                name = m.group(1)

            # 模式: "XXX评估程序中利集团" -> remove (garbage)
            if re.search(r'(?:评估程序|收款控制|测试流于)', name):
                continue

            # 跳过过短的名称
            if len(name) < 4:
                continue

            item['name'] = name
            cleaned.append(item)

        return json.dumps(cleaned, ensure_ascii=False) if cleaned else 'null'
    except:
        return tl_str


def main():
    input_file = '全量数据标注/STK_labeled_夏思远_2023110537_G02.xlsx'
    output_file = '全量数据标注/STK_labeled_夏思远_2023110537_G02.xlsx'

    print("正在读取标注结果...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    print(f"总行数: {ws.max_row - 1}")

    # 1. 应用 ann_related 修正
    print("\n=== 应用 ann_related 修正 ===")
    ann_corrected = 0
    for row_idx, correction in ann_related_corrections.items():
        if correction is None:
            continue

        ann_related, ann_year, ann_fin_flag, ann_fin_info, third_party_flag, third_party_list = correction

        if ann_related is not None:
            ws.cell(row=row_idx, column=2, value=ann_related)
        if ann_year is not None:
            ws.cell(row=row_idx, column=3, value=json.dumps(ann_year, ensure_ascii=False))
        elif ann_year is None and ann_related == 1:
            ws.cell(row=row_idx, column=3, value='null')
        if ann_fin_flag is not None:
            ws.cell(row=row_idx, column=4, value=ann_fin_flag if ann_fin_flag is not None else 'null')
        if ann_fin_info is not None:
            ws.cell(row=row_idx, column=5, value=json.dumps(ann_fin_info, ensure_ascii=False))
        elif ann_fin_info is None and ann_fin_flag == 1:
            ws.cell(row=row_idx, column=5, value='null')
        if third_party_flag is not None:
            ws.cell(row=row_idx, column=6, value=third_party_flag)
        if third_party_list is not None:
            ws.cell(row=row_idx, column=7, value=json.dumps(third_party_list, ensure_ascii=False))
        elif third_party_list is None and third_party_flag == 0:
            ws.cell(row=row_idx, column=7, value='null')

        ann_corrected += 1

    print(f"  已修正 {ann_corrected} 行")

    # 2. 移除审计机构误判的 third_party
    print("\n=== 修正 third_party 误判 ===")
    audit_rows = find_audit_firm_rows(ws)
    tp_removed = 0
    for r in audit_rows:
        ws.cell(row=r, column=6, value=0)
        ws.cell(row=r, column=7, value='null')
        tp_removed += 1

    print(f"  已移除 {tp_removed} 行审计机构误判")

    # 3. 清理第三方名称
    print("\n=== 清理第三方名称 ===")
    name_cleaned = 0
    for r in range(2, ws.max_row + 1):
        tp = ws.cell(row=r, column=6).value
        if tp == 1:
            tl = ws.cell(row=r, column=7).value
            cleaned = clean_third_party_list(tl)
            if cleaned != tl:
                ws.cell(row=r, column=7, value=cleaned)
                name_cleaned += 1
                # 如果清理后列表为空，改为0
                if cleaned == 'null':
                    ws.cell(row=r, column=6, value=0)

    print(f"  已清理 {name_cleaned} 行第三方名称")

    # 4. 统计修正后的结果
    ann_related_count = 0
    ann_fin_count = 0
    third_party_count = 0

    for r in range(2, ws.max_row + 1):
        ar = ws.cell(row=r, column=2).value
        af = ws.cell(row=r, column=4).value
        tp = ws.cell(row=r, column=6).value

        if ar == 1:
            ann_related_count += 1
        if af == 1:
            ann_fin_count += 1
        if tp == 1:
            third_party_count += 1

    print(f"\n=== 修正后统计 ===")
    print(f"ann_related=1: {ann_related_count}")
    print(f"ann_fin_flag=1: {ann_fin_count}")
    print(f"third_party_flag=1: {third_party_count}")

    wb.save(output_file)
    print(f"\n已保存: {output_file}")


if __name__ == '__main__':
    main()
