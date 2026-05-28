# -*- coding: utf-8 -*-
"""
全量财务舞弊标注 v2 - 改进版
大幅改进年报相关性检测和第三方识别
"""

import sys
import io
import re
import json
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ============================================================
# Task 1: 年报相关性判断 (v2 - 大幅改进)
# ============================================================

def extract_years(text):
    """从文本中提取与年报相关的年份"""
    years = set()

    # 精确匹配年报年份
    patterns = [
        r'(\d{4})年年度报告',
        r'(\d{4})年度报告',
        r'(\d{4})年年报',
        r'(\d{4})年报',
        r'(\d{4})年年度财务报告',
        r'(\d{4})年度财务报告',
        r'(\d{4})年年度财务报表',
        r'(\d{4})年度财务报表',
        r'(\d{4})年[年度]*报告.*(?:虚假|不实|不准确|遗漏|误导|未披露|披露不)',
    ]
    for p in patterns:
        for m in re.finditer(p, text):
            y = int(m.group(1))
            if 1990 <= y <= 2030:
                years.add(y)

    # 更宽泛的模式：如果已经有年报年份，再补充
    if years:
        extra_patterns = [
            r'(\d{4})年[年度]*(?:虚增|虚减|少计|多计)',
            r'(\d{4})年[年度]*(?:营业收入|净利润|利润总额|总资产|净资产)',
            r'(\d{4})年[年度]*(?:归母|归属于母公司)',
        ]
        for p in extra_patterns:
            for m in re.finditer(p, text):
                y = int(m.group(1))
                if 1990 <= y <= 2030:
                    years.add(y)

    return sorted(list(years)) if years else None


def is_annual_report_related(text):
    """
    判断是否年报相关 (v2)
    核心逻辑：违规行为是否导致年度报告中的信息存在错误或遗漏
    """
    # ========== 排除规则 ==========

    # 1. 仅涉及半年报/季报/季度报告
    has_semi = bool(re.search(r'半年[度]*报告|半年报|中期报告', text))
    has_quarter = bool(re.search(r'季度报告|季报|一季报|三季报', text))
    has_annual = bool(re.search(r'年度报告|年报(?!问询)', text))

    # 如果只提到半年报/季报，没有提到年报，则排除
    if (has_semi or has_quarter) and not has_annual:
        return False

    # 2. 窗口期交易（定期报告窗口期违规增持/减持）
    if re.search(r'窗口期|窗口.*(?:违规|买卖)|定期报告.*(?:前|内).*\d+.*(?:日|天).*(?:买卖|增[减持]|交易|卖出|买入|平仓)', text):
        # 但如果有虚假记载等，不排除
        if not re.search(r'(?:虚假|不实|不准确|虚增|虚减|少计|多计|遗漏|误导)', text):
            return False

    # 3. 年报问询函回复（只是回复问询，不是年报本身有问题）
    if re.search(r'(?:年报|年度报告)问询函.*回复|回复.*(?:年报|年度报告)问询函', text):
        # 但如果是对问询的回复中发现年报有问题，则不排除
        if not re.search(r'(?:虚假|不实|不准确|虚增|虚减|少计|多计|遗漏|误导)', text):
            return False

    # 4. 仅涉及内幕交易/操纵市场
    if re.search(r'内幕交易|操纵市场|操纵股价|操纵证券交易', text):
        if not re.search(r'(?:年报|年度报告).*(?:虚假|不实|不准确|遗漏|误导)', text):
            return False

    # 5. 仅涉及违规买卖股票（短线交易等）
    if re.search(r'(?:违规|违法).*买卖.*股票|短线交易', text):
        if not re.search(r'(?:年报|年度报告).*(?:虚假|不实|不准确|遗漏|误导)', text):
            return False

    # ========== 纳入规则 ==========

    # 规则A：明确提到年报存在虚假记载/重大遗漏/误导性陈述
    if re.search(r'(?:年度报告|年报).*(?:虚假[记]*载|重大遗漏|误导性陈述|不实|不准确|不完整)', text):
        return True
    if re.search(r'(?:虚假[记]*载|重大遗漏|误导性陈述|不实|不准确|不完整).*(?:年度报告|年报)', text):
        return True

    # 规则B：年报中未披露/披露不完整
    if re.search(r'(?:年度报告|年报).*(?:未披露|披露不|未.*(?:如实|充分|完整|及时).*披露)', text):
        return True

    # 规则C：导致年报数据错误的具体行为
    if re.search(r'(?:年度报告|年报).*(?:虚增|虚减|少计|多计)', text):
        return True
    if re.search(r'(?:虚增|虚减|少计|多计).*(?:年度报告|年报)', text):
        return True

    # 规则D：年报中的财务数据不真实
    if re.search(r'(?:年度报告|年报).*(?:财务数据|财务信息|财务报表).*(?:不真实|不准确|虚假)', text):
        return True

    # 规则E：招股说明书+年报虚假（欺诈上市）
    if re.search(r'招股说明书.*(?:虚假|不实|不准确)', text) and re.search(r'(?:年度报告|年报)', text):
        return True

    # 规则F：子公司造假导致母公司年报数据错误
    if re.search(r'(?:子公司|控股子公司).*(?:虚构|虚假|造假)', text) and re.search(r'(?:年度报告|年报)', text):
        if re.search(r'(?:虚增|虚减|少计|多计|收入|利润|资产)', text):
            return True

    # 规则G：年报中关联方/关联交易/担保披露问题
    if re.search(r'(?:年度报告|年报).*(?:关联[交易方]*|担保).*(?:未披露|披露不|遗漏)', text):
        return True
    if re.search(r'(?:关联[交易方]*|担保).*(?:未披露|披露不|遗漏).*(?:年度报告|年报)', text):
        return True

    # 规则H：直接描述"X年年报存在/存在以下违规/违法事实"
    if re.search(r'\d{4}年(?:年度报告|年报).*(?:存在|以下)', text):
        if re.search(r'(?:虚假|不实|不准确|遗漏|误导|未披露|违规)', text):
            return True

    # 规则I：违规行为描述中涉及年报中的具体会计处理
    if re.search(r'(?:年度报告|年报)', text):
        if re.search(r'(?:会计处理|会计估计|会计差错|会计政策).*(?:不恰当|不准确|违规|错误)', text):
            return True
        if re.search(r'(?:收入确认|成本结转|减值计提|折旧摊销).*(?:不恰当|不准确|违规|错误)', text):
            return True

    return False


# ============================================================
# Task 2: 财务信息与会计要素识别 (v2)
# ============================================================

# 非财务信息关键词（用于排除）
NON_FINANCIAL_PATTERNS = [
    r'(?:未|未及时|未按规定).*(?:披露|公告).*(?:关联[交易方]*|担保|诉讼|仲裁|重大合同|重大事项|股权变动|股份变动)',
    r'(?:关联[交易方]*|担保|诉讼|仲裁).*(?:未|未及时|未按规定).*(?:披露|公告|审议)',
    r'公司治理',
    r'内幕信息.*(?:知情人|登记|管理)',
    r'(?:董事|监事|高管).*(?:违规|违法).*买卖.*股票',
    r'股东.*(?:减持|增持).*未.*(?:披露|公告)',
    r'(?:实际控制人|控股股东).*披露不准确',
    r'同业竞争.*(?:未披露|违反)',
    r'承诺.*(?:未履行|违反)',
    r'信息披露.*(?:不及时|不完整).*(?!.*(?:收入|利润|资产|费用|成本))',
]

FINANCIAL_PATTERNS = [
    r'(?:虚增|虚减|少计|多计).*(?:收入|利润|资产|负债|费用|成本|净利润|营业收入|营业成本|净资产)',
    r'(?:收入|利润|资产|负债|费用|成本|净利润|营业收入|营业成本).*(?:虚增|虚减|少计|多计|虚假|不实|不准确)',
    r'(?:应收账款|存货|固定资产|在建工程|商誉|无形资产|货币资金|其他应收款).*(?:虚增|虚减|少计|多计)',
    r'(?:营业成本|管理费用|销售费用|财务费用|研发费用|所得税).*(?:少计|多计|虚增|虚减)',
    r'财务.*(?:虚假|不实|不准确|造假)',
    r'(?:会计处理|会计估计|会计差错|会计政策).*(?:不恰当|不准确|违规|错误)',
    r'(?:收入确认|成本结转|减值计提|折旧摊销|坏账准备).*(?:不恰当|不准确|违规|错误)',
    r'(?:合并报表|合并范围).*(?:不完整|不准确)',
    r'(?:毛利率|净利率|利润率).*(?:虚增|虚减)',
    r'财务报表.*(?:虚假|不实|不准确)',
    r'财务数据.*(?:虚假|不实|不准确|不真实)',
    r'(?:招股说明书|发行文件).*(?:虚假|不实|编造)',
]


def is_financial_info(text):
    """判断是否影响财务信息"""
    # 先检查明确的非财务模式
    for p in NON_FINANCIAL_PATTERNS:
        if re.search(p, text):
            has_fin = False
            for fp in FINANCIAL_PATTERNS:
                if re.search(fp, text):
                    has_fin = True
                    break
            if not has_fin:
                return False

    # 检查财务关键词
    for p in FINANCIAL_PATTERNS:
        if re.search(p, text):
            return True

    return False


# 会计要素映射
ELEMENT_KEYWORDS = {
    '资产': [r'资产', r'应收账款', r'存货', r'固定资产', r'在建工程', r'商誉',
             r'无形资产', r'货币资金', r'其他应收款', r'预付款项', r'长期股权投资',
             r'投资性房地产', r'使用权资产', r'开发支出', r'长期待摊费用',
             r'其他非流动资产', r'流动资产', r'非流动资产', r'总资产', r'净资产'],
    '负债': [r'负债', r'应付账款', r'短期借款', r'长期借款', r'预收款项',
             r'应付债券', r'其他应付款', r'一年内到期', r'长期应付款', r'预计负债',
             r'递延收益', r'递延所得税负债'],
    '所有者权益': [r'所有者权益', r'实收资本', r'股本', r'资本公积', r'盈余公积',
                   r'未分配利润', r'归属于母公司', r'少数股东权益', r'股东权益'],
    '收入': [r'收入', r'营业收入', r'主营业务收入', r'其他业务收入', r'销售收入',
             r'营业总收入', r'营收'],
    '费用': [r'费用', r'营业成本', r'主营业务成本', r'管理费用', r'销售费用',
             r'财务费用', r'研发费用', r'所得税费用', r'期间费用', r'营业总成本',
             r'成本', r'运杂费'],
    '利润': [r'利润', r'净利润', r'营业利润', r'利润总额', r'归母.*净利润',
             r'扣非.*净利润', r'综合收益', r'毛利'],
}


def identify_financial_elements(text, year):
    """识别特定年份受影响的会计要素"""
    elements = []
    for elem, patterns in ELEMENT_KEYWORDS.items():
        for p in patterns:
            if re.search(p, text):
                elements.append(elem)
                break
    return elements


def build_ann_fin_info(text, years):
    """构建ann_fin_info"""
    if not years:
        return None

    fin_info = []
    for year in years:
        elements = identify_financial_elements(text, year)
        if elements:
            fin_info.append({"year": year, "elements": elements})

    return fin_info if fin_info else None


# ============================================================
# Task 3: 第三方配合造假识别 (v2 - 改进)
# ============================================================

# 配合行为关键词
COLLUSION_KEYWORDS = [
    r'(?:配合|协助|参与).*(?:虚构|虚假|伪造|造假|舞弊)',
    r'签订.*(?:虚假|无商业实质).*(?:合同|协议)',
    r'(?:虚假|伪造).*(?:合同|协议|发票|凭证|验收|确认|单据)',
    r'(?:虚构|无真实).*(?:交易|业务|销售|采购)',
    r'资金.*(?:回流|循环|周转|出借)',
    r'出借.*(?:账户|银行账户|银行账号)',
    r'代持.*(?:股份|股权)',
    r'(?:配合|协助).*(?:虚增|虚减|隐瞒)',
    r'合谋|串通|共谋',
    r'配合.*(?:造假|舞弊)',
    r'提供.*(?:虚假|不实).*(?:凭证|发票|验收|确认|证明)',
    r'(?:无商业实质|无真实交易背景)',
    r'(?:配合|协助).*(?:资金|账务).*(?:处理|安排)',
]


def extract_third_party_names(text):
    """提取第三方名称 - 使用更精确的模式"""
    names = []

    # 模式1: "与XXX公司/企业签订虚假合同"
    p1 = re.finditer(r'(?:与|和)([一-龥A-Za-z]{4,40}(?:有限公司|股份有限公司|集团|合伙企业)(?:\([一-龥]*\))?)', text)
    for m in p1:
        name = m.group(1)
        # 排除上市公司自身和子公司
        if not re.search(r'(?:本公司|公司|发行人|上市公司|以下简称)', name):
            names.append(name)

    # 模式2: "XXX公司配合/协助/参与"
    p2 = re.finditer(r'([一-龥A-Za-z]{4,40}(?:有限公司|股份有限公司|集团|合伙企业)(?:\([一-龥]*\))?)(?:配合|协助|参与|签订|提供|出具|虚构)', text)
    for m in p2:
        name = m.group(1)
        if not re.search(r'(?:本公司|公司|发行人|上市公司|以下简称)', name):
            if name not in names:
                names.append(name)

    # 模式3: 从"通过XXX公司"模式中提取
    p3 = re.finditer(r'通过(?:其|控制的)?([一-龥A-Za-z]{4,40}(?:有限公司|股份有限公司|合伙企业))', text)
    for m in p3:
        name = m.group(1)
        if not re.search(r'(?:本公司|子公司|控股子公司|全资子公司)', name):
            if name not in names:
                names.append(name)

    return names if names else None


def get_third_party_type(name, activity_text=""):
    """根据名称和Activity上下文推断第三方类型"""
    # 先根据名称判断
    if re.search(r'银行|支行|分行|信托|基金|证券公司|保险|金融', name):
        return '银行/金融机构'
    if re.search(r'会计师事务所|审计', name):
        return '会计师事务所'
    if re.search(r'评估|资产评估|估值', name):
        return '评估机构'
    if re.search(r'券商|保荐|承销|投行', name):
        return '券商/保荐机构'

    # 根据Activity上下文判断
    # 客户：签订销售合同/买方/采购方
    if re.search(r'(?:客户|买方|采购方|下游|销售对象|购买方)', activity_text):
        return '客户'
    if re.search(r'(?:签订|提供).*(?:销售|卖|出售).*(?:合同|协议|确认)', activity_text):
        return '客户'

    # 供应商：签订采购合同/卖方/供货方
    if re.search(r'(?:供应商|供货方|上游|采购对象|卖方)', activity_text):
        return '供应商'
    if re.search(r'(?:签订|提供).*(?:采购|买|购买).*(?:合同|协议)', activity_text):
        return '供应商'

    # 自然人
    if re.search(r'自然人|个人|代持', name):
        return '自然人'

    # 兜底
    return '其他企业'


def get_collusion_role(text):
    """提取配合行为描述"""
    for p in COLLUSION_KEYWORDS:
        m = re.search(p, text)
        if m:
            role = m.group(0)
            # 截取合理长度
            if len(role) > 80:
                role = role[:80]
            return role
    return '配合财务舞弊'


def is_excluded_entity(name, text):
    """判断是否应排除的实体（非第三方）"""
    # 控股股东/实控人
    if re.search(r'控股股东|实际控制人|实控人', name):
        return True
    # 子公司
    if re.search(r'子公司|全资子公司|控股子公司', name):
        return True
    # 仅以"控股股东""实控人"等描述性词语开头的名称
    if re.search(r'^(?:控股股东|实际控制人|实控人)', name):
        return True
    # 被收购标的公司（在重大资产重组上下文中）
    if re.search(r'(?:收购|购买|重组).*(?:股权|资产)', text):
        if re.search(r'(?:标的|目标|被收购)', text):
            return True
    return False


def clean_entity_name(name):
    """清理实体名称中的垃圾文本"""
    # 移除常见的垃圾前缀
    garbage_prefixes = [
        r'^评估明显不到位',
        r'^管理层怠于履行向',
        r'^交易资金系由[^授权]*授权[^负责]*全权负责的',
        r'^发行股份及支付现金的方式购买',
        r'^虚构与',
        r'^与',
        r'^通过(?:其|控制的)?',
        r'^有效的内部控制程序发现并识别',
        r'^收款控制测试流于形式',
        r'^代持股份',
    ]
    for p in garbage_prefixes:
        m = re.search(p, name)
        if m:
            name = name[m.end():]

    # 提取实际公司名称
    m = re.search(r'([一-龥A-Za-z]{2,40}(?:有限公司|股份有限公司|合伙企业|集团)(?:\([一-龥]*\))?)', name)
    if m:
        return m.group(1)

    return name.strip()


def identify_third_party(text):
    """识别第三方配合造假"""
    # 检查是否有配合造假关键词
    has_collusion = False
    for p in COLLUSION_KEYWORDS:
        if re.search(p, text):
            has_collusion = True
            break

    if not has_collusion:
        return None

    # 提取第三方名称
    names = extract_third_party_names(text)
    if not names:
        return None

    # 清理和过滤
    cleaned = []
    for name in names:
        # 清理名称
        name = clean_entity_name(name)
        if not name or len(name) < 3:
            continue
        # 排除非第三方实体
        if is_excluded_entity(name, text):
            continue
        # 去重
        if name not in [c['name'] for c in cleaned]:
            cleaned.append({
                "name": name,
                "type": get_third_party_type(name, text),
                "role": get_collusion_role(text),
            })

    return cleaned if cleaned else None

    # 构建第三方列表
    third_parties = []
    role = get_collusion_role(text)
    for name in names[:5]:  # 最多5个
        third_parties.append({
            "name": name,
            "type": get_third_party_type(name),
            "role": role,
        })

    return third_parties


# ============================================================
# 主处理流程
# ============================================================

def process_row(text):
    """处理单行数据"""
    if not text or not text.strip():
        return {
            'ann_related': 0,
            'ann_year': None,
            'ann_fin_flag': None,
            'ann_fin_info': None,
            'third_party_flag': None,
            'third_party_list': None,
        }

    # Task 1
    ann_related = 1 if is_annual_report_related(text) else 0
    ann_year = extract_years(text) if ann_related == 1 else None

    # Task 2
    if ann_related == 1:
        if is_financial_info(text):
            ann_fin_flag = 1
            ann_fin_info = build_ann_fin_info(text, ann_year)
        else:
            ann_fin_flag = 0
            ann_fin_info = None
    else:
        ann_fin_flag = None
        ann_fin_info = None

    # Task 3
    if ann_related == 1 and ann_fin_flag == 1:
        third_party_list = identify_third_party(text)
        if third_party_list:
            third_party_flag = 1
        else:
            third_party_flag = 0
            third_party_list = None
    else:
        third_party_flag = None
        third_party_list = None

    return {
        'ann_related': ann_related,
        'ann_year': ann_year,
        'ann_fin_flag': ann_fin_flag,
        'ann_fin_info': ann_fin_info,
        'third_party_flag': third_party_flag,
        'third_party_list': third_party_list,
    }


def main():
    input_file = 'STK_Violation_Main_第2组.xlsx'
    output_file = '全量数据标注/STK_labeled_夏思远_2023110537_G02.xlsx'

    print("正在读取数据...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # 创建输出工作簿
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active

    # 写入表头
    headers = ['Activity', 'ann_related', 'ann_year', 'ann_fin_flag',
               'ann_fin_info', 'third_party_flag', 'third_party_list']
    for col, h in enumerate(headers, 1):
        out_ws.cell(row=1, column=col, value=h)

    # 处理数据
    total = 0
    ann_related_count = 0
    ann_fin_count = 0
    third_party_count = 0

    print(f"开始处理数据...")
    for row_idx in range(4, ws.max_row + 1):
        activity = ws.cell(row=row_idx, column=17).value
        if not activity or activity == '没有单位' or not str(activity).strip():
            continue

        total += 1
        result = process_row(str(activity))

        # 写入输出
        out_row = total + 1
        out_ws.cell(row=out_row, column=1, value=str(activity))
        out_ws.cell(row=out_row, column=2, value=result['ann_related'])
        out_ws.cell(row=out_row, column=3, value=json.dumps(result['ann_year'], ensure_ascii=False) if result['ann_year'] else 'null')
        out_ws.cell(row=out_row, column=4, value=result['ann_fin_flag'] if result['ann_fin_flag'] is not None else 'null')
        out_ws.cell(row=out_row, column=5, value=json.dumps(result['ann_fin_info'], ensure_ascii=False) if result['ann_fin_info'] else 'null')
        out_ws.cell(row=out_row, column=6, value=result['third_party_flag'] if result['third_party_flag'] is not None else 'null')
        out_ws.cell(row=out_row, column=7, value=json.dumps(result['third_party_list'], ensure_ascii=False) if result['third_party_list'] else 'null')

        # 统计
        if result['ann_related'] == 1:
            ann_related_count += 1
        if result['ann_fin_flag'] == 1:
            ann_fin_count += 1
        if result['third_party_flag'] == 1:
            third_party_count += 1

        if total % 500 == 0:
            print(f"  已处理 {total} 行...")

    out_wb.save(output_file)
    print(f"\n=== 处理完成 ===")
    print(f"总数据行: {total}")
    print(f"ann_related=1: {ann_related_count}")
    print(f"ann_fin_flag=1: {ann_fin_count}")
    print(f"third_party_flag=1: {third_party_count}")
    print(f"\n已保存: {output_file}")


if __name__ == '__main__':
    main()
