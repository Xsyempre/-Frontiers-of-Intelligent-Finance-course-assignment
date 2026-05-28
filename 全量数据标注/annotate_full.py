# -*- coding: utf-8 -*-
"""
全量财务舞弊标注 - 第2组数据
"""

import sys
import io
import re
import json
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ============================================================
# Task 1: 年报相关性判断
# ============================================================

# 年报相关关键词
ANNUAL_REPORT_PATTERNS = [
    r'\d{4}年年度报告',
    r'\d{4}年度报告',
    r'\d{4}年报',
    r'年度报告[^中]*虚假',
    r'年度报告[^中]*重大遗漏',
    r'年度报告[^中]*误导',
    r'年度报告[^中]*不实',
    r'年度报告[^中]*不准确',
    r'年度报告[^中]*不完整',
    r'年度报告[^中]*未披露',
    r'年度报告[^中]*披露不',
    r'年报[^中]*虚假',
    r'年报[^中]*重大遗漏',
    r'年报[^中]*不实',
    r'年报[^中]*不准确',
]

# 排除模式：仅涉及季报/半年报/季度报告
EXCLUDE_PATTERNS = [
    r'仅.*半年报',
    r'仅.*季度报告',
    r'仅.*季报',
    r'半年报.*未.*年报.*已',
    r'季度报告.*未.*年报.*已',
]

# 非年报相关模式
NON_ANNUAL_PATTERNS = [
    r'违规买卖股票',
    r'短线交易',
    r'内幕交易',
    r'操纵市场',
    r'操纵股价',
    r'证券账户',
    r'交易股票',
    r'减持.*未.*披露',
    r'增持.*未.*披露',
    r'权益变动.*未.*披露',
    r'安全事故',
    r'安全生产',
    r'产品质量',
    r'化妆品',
    r'药品.*质量',
    r'环保.*处罚',
    r'税务.*处罚',
]

def extract_years(text):
    """从文本中提取年份"""
    years = set()
    # 优先匹配年报相关年份
    patterns = [
        r'(\d{4})年年度报告',
        r'(\d{4})年度报告',
        r'(\d{4})年报',
        r'(\d{4})年[年度]*财务报告',
        r'(\d{4})年[年度]*财务报表',
    ]
    for p in patterns:
        for m in re.finditer(p, text):
            y = int(m.group(1))
            if 1990 <= y <= 2030:
                years.add(y)

    # 如果没找到，尝试更宽泛的模式
    if not years:
        broad_patterns = [
            r'(\d{4})年[年度]*(?:虚增|虚减|少计|多计)',
            r'(\d{4})年[年度]*(?:营业收入|利润|资产|负债|费用|成本)',
            r'(\d{4})年(?:的)?(?:年报|年度报告)',
        ]
        for p in broad_patterns:
            for m in re.finditer(p, text):
                y = int(m.group(1))
                if 1990 <= y <= 2030:
                    years.add(y)

    return sorted(list(years)) if years else None


def is_annual_report_related(text):
    """判断是否年报相关"""
    # 先检查排除模式
    for p in EXCLUDE_PATTERNS:
        if re.search(p, text):
            return False

    # 检查非年报模式
    is_non_annual = False
    for p in NON_ANNUAL_PATTERNS:
        if re.search(p, text):
            is_non_annual = True
            break

    # 检查年报相关关键词
    has_annual_ref = False
    for p in ANNUAL_REPORT_PATTERNS:
        if re.search(p, text):
            has_annual_ref = True
            break

    if has_annual_ref and not is_non_annual:
        return True

    # 额外检查：如果提到"年报"或"年度报告"并且有虚假/遗漏/误导等词
    if re.search(r'(?:年报|年度报告)', text) and re.search(r'(?:虚假|遗漏|误导|不实|不准确|不完整|未披露|披露不)', text):
        if not is_non_annual:
            return True

    return False


# ============================================================
# Task 2: 财务信息与会计要素识别
# ============================================================

# 财务信息关键词
FINANCIAL_KEYWORDS = [
    r'虚增.*(?:收入|利润|资产|营业收入|净利润|净资产)',
    r'虚减.*(?:收入|利润|资产|费用|成本)',
    r'少计.*(?:费用|成本|负债|损失)',
    r'多计.*(?:收入|利润|资产)',
    r'(?:收入|利润|资产|费用|成本|负债).*虚假',
    r'(?:收入|利润|资产|费用|成本|负债).*不实',
    r'(?:收入|利润|资产|费用|成本|负债).*不准确',
    r'财务.*(?:虚假|不实|不准确)',
    r'(?:营业收入|净利润|净资产|总资产).*(?:虚增|虚减|少计|多计)',
    r'会计.*(?:差错|处理)',
    r'(?:应收账款|存货|固定资产|在建工程|商誉|无形资产).*(?:虚增|虚减|少计|多计)',
    r'(?:营业成本|管理费用|销售费用|财务费用|研发费用).*(?:少计|多计|虚增|虚减)',
    r'(?:毛利率|净利率|利润)',
    r'(?:折旧|摊销|减值).*(?:少计|多计)',
    r'合并报表',
    r'财务报表',
    r'财务数据',
]

# 非财务信息关键词
NON_FINANCIAL_KEYWORDS = [
    r'未披露.*(?:关联|担保|诉讼|仲裁|重大合同)',
    r'披露不.*(?:完整|准确|及时).*(?:关联|担保|诉讼)',
    r'公司治理',
    r'内幕信息',
    r'知情人',
    r'重大事项.*未.*披露',
    r'关联交易.*未.*披露',
    r'关联方.*未.*披露',
    r'对外担保.*未.*披露',
    r'诉讼.*未.*披露',
    r'仲裁.*未.*披露',
    r'违规担保',
    r'资金占用',
    r'股份.*(?:减持|增持).*未.*披露',
]

# 会计要素映射
ELEMENT_KEYWORDS = {
    '资产': [r'资产', r'应收账款', r'存货', r'固定资产', r'在建工程', r'商誉', r'无形资产',
             r'货币资金', r'其他应收款', r'预付款项', r'长期股权投资', r'投资性房地产',
             r'使用权资产', r'生物资产', r'油气资产', r'开发支出', r'长期待摊费用',
             r'其他非流动资产', r'流动资产', r'非流动资产', r'总资产', r'净资产'],
    '负债': [r'负债', r'应付账款', r'短期借款', r'长期借款', r'预收款项', r'应付债券',
             r'其他应付款', r'一年内到期的非流动负债', r'长期应付款', r'预计负债',
             r'递延收益', r'递延所得税负债', r'其他流动负债', r'其他非流动负债'],
    '所有者权益': [r'所有者权益', r'实收资本', r'股本', r'资本公积', r'盈余公积',
                   r'未分配利润', r'归属于母公司', r'少数股东权益', r'股东权益'],
    '收入': [r'收入', r'营业收入', r'主营业务收入', r'其他业务收入', r'销售收入',
             r'营业收入', r'营业总收入'],
    '费用': [r'费用', r'营业成本', r'管理费用', r'销售费用', r'财务费用', r'研发费用',
             r'所得税费用', r'期间费用', r'主营业务成本', r'营业总成本', r'成本'],
    '利润': [r'利润', r'净利润', r'营业利润', r'利润总额', r'归属于母公司.*净利润',
             r'扣非.*净利润', r'综合收益', r'毛利', r'毛利率'],
}


def identify_financial_elements(text, year):
    """识别特定年份受影响的会计要素"""
    elements = []
    for elem, patterns in ELEMENT_KEYWORDS.items():
        for p in patterns:
            if re.search(p, text):
                elements.append(elem)
                break
    return elements


def is_financial_info(text):
    """判断是否影响财务信息"""
    # 先检查非财务模式
    for p in NON_FINANCIAL_KEYWORDS:
        if re.search(p, text):
            # 如果同时有财务关键词，仍算财务
            has_fin = False
            for fp in FINANCIAL_KEYWORDS:
                if re.search(fp, text):
                    has_fin = True
                    break
            if not has_fin:
                return False

    # 检查财务关键词
    for p in FINANCIAL_KEYWORDS:
        if re.search(p, text):
            return True

    return False


def build_ann_fin_info(text, years):
    """构建ann_fin_info"""
    if not years:
        return None

    fin_info = []
    for year in years:
        elements = identify_financial_elements(text, year)
        if elements:
            fin_info.append({"year": year, "elements": elements})

    return fin_info if fin_info else None


# ============================================================
# Task 3: 第三方配合造假识别
# ============================================================

# 第三方类型关键词
THIRD_PARTY_TYPES = {
    '客户': [r'客户', r'买方', r'采购方', r'购买方'],
    '供应商': [r'供应商', r'卖方', r'供货方', r'供方'],
    '银行/金融机构': [r'银行', r'金融机构', r'信托', r'基金', r'证券公司'],
    '券商/保荐机构': [r'券商', r'保荐', r'承销', r'投行'],
    '会计师事务所': [r'会计师事务所', r'审计', r'审计机构'],
    '评估机构': [r'评估', r'资产评估', r'评估机构'],
    '自然人': [r'自然人', r'个人'],
    '其他企业': [r'公司', r'企业', r'有限', r'合伙'],
}

# 配合行为关键词
COLLUSION_KEYWORDS = [
    r'配合.*(?:虚构|虚假|伪造)',
    r'签订.*虚假.*合同',
    r'虚假.*(?:交易|业务|合同)',
    r'资金.*(?:回流|循环|周转)',
    r'出借.*(?:账户|银行账户)',
    r'代持.*股份',
    r'协助.*(?:虚构|虚假|伪造|隐瞒)',
    r'合谋',
    r'串通',
    r'共谋',
    r'配合.*造假',
    r'配合.*舞弊',
    r'配合.*虚增',
    r'提供.*虚假.*(?:凭证|发票|验收|确认)',
    r'伪造.*(?:合同|发票|凭证|单据)',
    r'虚构.*(?:交易|业务|合同|销售|采购)',
    r'无商业实质',
    r'无真实交易',
]

# 排除的第三方（不算配合造假）
EXCLUDE_THIRD_PARTY = [
    r'审计.*(?:意见|报告)',
    r'出具.*审计',
    r'正常.*(?:交易|业务)',
    r'配合.*调查',
    r'配合.*监管',
]


def identify_third_party(text):
    """识别第三方配合造假"""
    # 先检查排除模式
    for p in EXCLUDE_THIRD_PARTY:
        if re.search(p, text):
            # 但如果同时有配合造假关键词，仍算
            has_collusion = False
            for cp in COLLUSION_KEYWORDS:
                if re.search(cp, text):
                    has_collusion = True
                    break
            if not has_collusion:
                return None

    # 检查配合造假关键词
    has_collusion = False
    for p in COLLUSION_KEYWORDS:
        if re.search(p, text):
            has_collusion = True
            break

    if not has_collusion:
        return None

    # 尝试提取第三方名称和类型
    third_parties = []

    # 模式1: "与XXX公司签订虚假合同"
    patterns = [
        r'(?:与|和)([一-龥]{2,20}(?:公司|企业|银行|事务所|机构|集团|合伙)(?:\([一-龥]*\))?)',
        r'([一-龥]{2,20}(?:公司|企业|银行|事务所|机构|集团|合伙)(?:\([一-龥]*\))?)(?:配合|协助|参与|签订|提供|出具)',
    ]

    for p in patterns:
        for m in re.finditer(p, text):
            name = m.group(1)
            # 确定类型
            ptype = '其他企业'
            for tname, tpatterns in THIRD_PARTY_TYPES.items():
                for tp in tpatterns:
                    if re.search(tp, name):
                        ptype = tname
                        break
                if ptype != '其他企业':
                    break

            # 生成角色描述
            role = '配合财务舞弊'
            for cp in COLLUSION_KEYWORDS:
                cm = re.search(cp, text)
                if cm:
                    role = cm.group(0)[:50]
                    break

            third_parties.append({
                "name": name,
                "type": ptype,
                "role": role
            })

    return third_parties if third_parties else None


# ============================================================
# 主处理流程
# ============================================================

def process_row(text):
    """处理单行数据"""
    if not text or not text.strip():
        return {
            'ann_related': 0,
            'ann_year': None,
            'ann_fin_flag': None,
            'ann_fin_info': None,
            'third_party_flag': None,
            'third_party_list': None,
        }

    # Task 1
    ann_related = 1 if is_annual_report_related(text) else 0
    ann_year = extract_years(text) if ann_related == 1 else None

    # Task 2
    if ann_related == 1:
        if is_financial_info(text):
            ann_fin_flag = 1
            ann_fin_info = build_ann_fin_info(text, ann_year)
        else:
            ann_fin_flag = 0
            ann_fin_info = None
    else:
        ann_fin_flag = None
        ann_fin_info = None

    # Task 3
    if ann_related == 1 and ann_fin_flag == 1:
        third_party_list = identify_third_party(text)
        if third_party_list:
            third_party_flag = 1
        else:
            third_party_flag = 0
            third_party_list = None
    else:
        third_party_flag = None
        third_party_list = None

    return {
        'ann_related': ann_related,
        'ann_year': ann_year,
        'ann_fin_flag': ann_fin_flag,
        'ann_fin_info': ann_fin_info,
        'third_party_flag': third_party_flag,
        'third_party_list': third_party_list,
    }


def main():
    input_file = 'STK_Violation_Main_第2组.xlsx'
    output_file = '全量数据标注/STK_labeled_夏思远_2023110537_G02.xlsx'

    print("正在读取数据...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # 创建输出工作簿
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active

    # 写入表头
    headers = ['Activity', 'ann_related', 'ann_year', 'ann_fin_flag',
               'ann_fin_info', 'third_party_flag', 'third_party_list']
    for col, h in enumerate(headers, 1):
        out_ws.cell(row=1, column=col, value=h)

    # 处理数据（从第4行开始，跳过前3行表头/单位行）
    total = 0
    ann_related_count = 0
    ann_fin_count = 0
    third_party_count = 0

    print(f"开始处理数据...")
    for row_idx in range(4, ws.max_row + 1):
        activity = ws.cell(row=row_idx, column=17).value
        if not activity or activity == '没有单位' or not str(activity).strip():
            continue

        total += 1
        result = process_row(str(activity))

        # 写入输出
        out_row = total + 1  # +1 for header
        out_ws.cell(row=out_row, column=1, value=str(activity))
        out_ws.cell(row=out_row, column=2, value=result['ann_related'])
        out_ws.cell(row=out_row, column=3, value=json.dumps(result['ann_year'], ensure_ascii=False) if result['ann_year'] else 'null')
        out_ws.cell(row=out_row, column=4, value=result['ann_fin_flag'] if result['ann_fin_flag'] is not None else 'null')
        out_ws.cell(row=out_row, column=5, value=json.dumps(result['ann_fin_info'], ensure_ascii=False) if result['ann_fin_info'] else 'null')
        out_ws.cell(row=out_row, column=6, value=result['third_party_flag'] if result['third_party_flag'] is not None else 'null')
        out_ws.cell(row=out_row, column=7, value=json.dumps(result['third_party_list'], ensure_ascii=False) if result['third_party_list'] else 'null')

        # 统计
        if result['ann_related'] == 1:
            ann_related_count += 1
        if result['ann_fin_flag'] == 1:
            ann_fin_count += 1
        if result['third_party_flag'] == 1:
            third_party_count += 1

        if total % 500 == 0:
            print(f"  已处理 {total} 行...")

    out_wb.save(output_file)
    print(f"\n=== 处理完成 ===")
    print(f"总数据行: {total}")
    print(f"ann_related=1: {ann_related_count}")
    print(f"ann_fin_flag=1: {ann_fin_count}")
    print(f"third_party_flag=1: {third_party_count}")
    print(f"\n已保存: {output_file}")


if __name__ == '__main__':
    main()
