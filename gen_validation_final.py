"""
严格按照 具体要求03——生成交付物三_人工标注与模型对比表（validation.xlsx）.md 的要求生成 validation.xlsx
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========== 读取数据 ==========
benchmark = pd.read_excel('benchmark_G02.xlsx')
deepseek = pd.read_excel('DeepSeek/deepseek_ver.xlsx')
minimax = pd.read_excel('MiniMax/minimax_ver.xlsx')
mimo = pd.read_excel('Mimo/mimo_ver.xlsx')

# ========== 按 SampleSeq 对齐，提取需要的字段 ==========
bm = benchmark[['SampleSeq', 'Activity', 'ann_related', 'ann_fin_flag', 'third_party_flag']].copy()
bm.columns = ['SampleSeq', 'Activity', 'human_ann', 'human_fin', 'human_tp']

dp = deepseek[['SampleSeq', 'ann_related', 'ann_fin_flag', 'third_party_flag']].copy()
dp.columns = ['SampleSeq', 'deep_ann', 'deep_fin', 'deep_tp']

mn = minimax[['SampleSeq', 'ann_related', 'ann_fin_flag', 'third_party_flag']].copy()
mn.columns = ['SampleSeq', 'mini_ann', 'mini_fin', 'mini_tp']

mm = mimo[['SampleSeq', 'ann_related', 'ann_fin_flag', 'third_party_flag']].copy()
mm.columns = ['SampleSeq', 'mimo_ann', 'mimo_fin', 'mimo_tp']

df = bm.merge(dp, on='SampleSeq').merge(mn, on='SampleSeq').merge(mm, on='SampleSeq')
print(f"样本数: {len(df)}")

# ========== 辅助函数 ==========
def to_val(v):
    """将 NaN/None 转为字符串 'null'，0/1 转为 '0'/'1'，其他转为字符串"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 'null'
    if v in (0, 1):
        return str(int(v))
    return str(v)

def calc_cm(human_val, model_val):
    """返回 (TP, FP, FN, TN) 元组。human_val 不会为 null（有效子集），model_val 为 null 时按 0 处理。"""
    h = int(human_val)
    m = int(model_val) if not (model_val is None or (isinstance(model_val, float) and pd.isna(model_val))) else 0
    if h == 1 and m == 1:
        return (1, 0, 0, 0)
    elif h == 0 and m == 1:
        return (0, 1, 0, 0)
    elif h == 1 and m == 0:
        return (0, 0, 1, 0)
    elif h == 0 and m == 0:
        return (0, 0, 0, 1)
    return (None, None, None, None)

def calc_f1(tp, fp, fn):
    if tp is None or (tp + fp + fn) == 0:
        return None
    prec = tp / (tp + fp) if (tp + fp) > 0 else 0
    rec = tp / (tp + fn) if (tp + fn) > 0 else 0
    if prec + rec == 0:
        return 0
    return 2 * prec * rec / (prec + rec)

# ========== 计算每行的 TP/FP/FN/TN ==========
models = [('deep', 'DeepSeek'), ('mini', 'MiniMax'), ('mimo', 'Mimo')]

# 为每行计算 confusion matrix
for model_prefix, _ in models:
    ann_tp, ann_fp, ann_fn, ann_tn = [], [], [], []
    fin_tp, fin_fp, fin_fn, fin_tn = [], [], [], []
    tp_tp, tp_fp, tp_fn, tp_tn = [], [], [], []

    for i in range(len(df)):
        row = df.iloc[i]

        # ann_related (所有行)
        t, f, fn, tn = calc_cm(row['human_ann'], row[f'{model_prefix}_ann'])
        ann_tp.append(t); ann_fp.append(f); ann_fn.append(fn); ann_tn.append(tn)

        # ann_fin_flag (仅 human_ann == 1 的行)
        if row['human_ann'] == 1:
            t, f, fn, tn = calc_cm(row['human_fin'], row[f'{model_prefix}_fin'])
        else:
            t, f, fn, tn = None, None, None, None
        fin_tp.append(t); fin_fp.append(f); fin_fn.append(fn); fin_tn.append(tn)

        # third_party_flag (仅 human_ann == 1 且 human_fin == 1 的行)
        if row['human_ann'] == 1 and row['human_fin'] == 1:
            t, f, fn, tn = calc_cm(row['human_tp'], row[f'{model_prefix}_tp'])
        else:
            t, f, fn, tn = None, None, None, None
        tp_tp.append(t); tp_fp.append(f); tp_fn.append(fn); tp_tn.append(tn)

    df[f'{model_prefix}_ann_tp'] = ann_tp
    df[f'{model_prefix}_ann_fp'] = ann_fp
    df[f'{model_prefix}_ann_fn'] = ann_fn
    df[f'{model_prefix}_ann_tn'] = ann_tn
    df[f'{model_prefix}_fin_tp'] = fin_tp
    df[f'{model_prefix}_fin_fp'] = fin_fp
    df[f'{model_prefix}_fin_fn'] = fin_fn
    df[f'{model_prefix}_fin_tn'] = fin_tn
    df[f'{model_prefix}_tp_tp'] = tp_tp
    df[f'{model_prefix}_tp_fp'] = tp_fp
    df[f'{model_prefix}_tp_fn'] = tp_fn
    df[f'{model_prefix}_tp_tn'] = tp_tn

# ========== 计算各模型各字段的 F1 ==========
results = {}
for model_prefix, model_name in models:
    # ann_related F1 (全部行)
    tp = df[f'{model_prefix}_ann_tp'].sum()
    fp = df[f'{model_prefix}_ann_fp'].sum()
    fn = df[f'{model_prefix}_ann_fn'].sum()
    f1_ann = calc_f1(tp, fp, fn)

    # ann_fin_flag F1 (human_ann == 1 子集)
    sub = df[df['human_ann'] == 1]
    tp = sub[f'{model_prefix}_fin_tp'].sum()
    fp = sub[f'{model_prefix}_fin_fp'].sum()
    fn = sub[f'{model_prefix}_fin_fn'].sum()
    f1_fin = calc_f1(tp, fp, fn)

    # third_party_flag F1 (human_ann == 1 and human_fin == 1 子集)
    sub2 = df[(df['human_ann'] == 1) & (df['human_fin'] == 1)]
    tp = sub2[f'{model_prefix}_tp_tp'].sum()
    fp = sub2[f'{model_prefix}_tp_fp'].sum()
    fn = sub2[f'{model_prefix}_tp_fn'].sum()
    f1_tp_val = calc_f1(tp, fp, fn)

    valid = [v for v in [f1_ann, f1_fin, f1_tp_val] if v is not None]
    avg = sum(valid) / len(valid) if valid else None

    results[model_prefix] = {
        'name': model_name,
        'f1_ann': f1_ann, 'f1_fin': f1_fin, 'f1_tp': f1_tp_val, 'avg': avg,
        'n_fin': len(sub), 'n_tp': len(sub2)
    }

print("\n=== F1 Score ===")
for mk, rv in results.items():
    print(f"{rv['name']}: ann={rv['f1_ann']:.4f} | fin={rv['f1_fin']} | tp={rv['f1_tp']} | avg={rv['avg']:.4f} | n_fin={rv['n_fin']}, n_tp={rv['n_tp']}")

# ========== 生成 Excel ==========
wb = Workbook()

# 样式定义
H_FILL = PatternFill("solid", fgColor="1A1A18")
S_FILL = PatternFill("solid", fgColor="378ADD")
L_FILL = PatternFill("solid", fgColor="F1EFE8")
W_FILL = PatternFill("solid", fgColor="FFFFFF")
G_FILL = PatternFill("solid", fgColor="E8E6DE")
GR_FILL = PatternFill("solid", fgColor="EAF3DE")
HU_FILL = PatternFill("solid", fgColor="FAEEDA")
ALT_FILL = PatternFill("solid", fgColor="F8F7F3")

THIN = Border(left=Side(style='thin', color="1A1A18"),
              right=Side(style='thin', color="1A1A18"),
              top=Side(style='thin', color="1A1A18"),
              bottom=Side(style='thin', color="1A1A18"))

def cs(cell, val, bold=False, fc="1A1A18", fill=None, size=10, wrap=False, align='center'):
    cell.value = val
    cell.font = Font(name='Courier New', size=size, bold=bold, color=fc)
    if fill: cell.fill = fill
    cell.alignment = Alignment(wrap_text=wrap, horizontal=align, vertical='center')
    cell.border = THIN

# ============================================================
# Sheet 1: 模型对比总览（类似作业说明中的示例格式）
# ============================================================
ws1 = wb.active
ws1.title = "模型对比总览"

# 标题
ws1.merge_cells('A1:E1')
cs(ws1['A1'], '模型横向对比表（5%样本）', bold=True, fc="FFFFFF", fill=H_FILL, size=13, align='center')
ws1.row_dimensions[1].height = 30

# 说明行
ws1.merge_cells('A2:E2')
cs(ws1['A2'], 'ann_fin_flag仅在ann_related=1的子样本中计算；third_party_flag仅在ann_related=1且ann_fin_flag=1的子样本中计算。',
   fc="4A4A46", fill=L_FILL, size=9, align='left', wrap=True)
ws1.row_dimensions[2].height = 22

# Part I 小标题
ws1.merge_cells('A3:E3')
cs(ws1['A3'], 'Part I：模型F1 Score横向对比汇总', bold=True, fc="FFFFFF", fill=S_FILL, size=11, align='left')
ws1.row_dimensions[3].height = 22

# 表头
for c, h in enumerate(['评估字段', '人工标注(基准)', 'DeepSeek', 'MiniMax', 'Mimo'], 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.fill = G_FILL if c == 1 else HU_FILL if c == 2 else W_FILL
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN
ws1.row_dimensions[4].height = 28

# 数据行
for row_i, (field, key) in enumerate([
    ('ann_related（是否影响年报）', 'f1_ann'),
    ('ann_fin_flag（是否影响财务信息）', 'f1_fin'),
    ('third_party_flag（是否有第三方）', 'f1_tp'),
    ('平均F1', 'avg'),
], 5):
    vals = [results[m][key] for m in ['deep', 'mini', 'mimo']]
    fmtd = ['N/A' if v is None else f'{v:.4f}' for v in vals]
    valid = [(v, m) for v, m in zip(vals, ['deep', 'mini', 'mimo']) if v is not None]
    best = max(valid, key=lambda x: x[0])[1] if valid else None

    row_data = [field, '—'] + fmtd
    rf = W_FILL if row_i % 2 == 0 else ALT_FILL

    for c, v in enumerate(row_data, 1):
        cell = ws1.cell(row=row_i, column=c, value=v)
        cell.fill = rf
        mk = ['deep', 'mini', 'mimo'][c - 3] if c >= 3 else None
        if mk == best and key == 'avg':
            cell.fill = GR_FILL
            cell.font = Font(name='Courier New', size=10, bold=True, color="3B6D11")
            cell.value = fmtd[c - 3] + ' ✓'
        else:
            cell.font = Font(name='Courier New', size=10, color="1A1A18")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN
    ws1.row_dimensions[row_i].height = 20

# 有效样本说明
note_row = 9
ws1.merge_cells(f'A{note_row}:E{note_row}')
note = (f'注：ann_fin_flag有效样本数={results["deep"]["n_fin"]}条；'
        f'third_party_flag有效样本数={results["deep"]["n_tp"]}条')
cs(ws1[f'A{note_row}'], note, fc="4A4A46", fill=L_FILL, size=9, align='left')

ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 14

# ============================================================
# Sheet 2: 详细逐条记录
# ============================================================
ws2 = wb.create_sheet("详细逐条记录")

ws2.merge_cells('A1:AW1')
cs(ws2['A1'], 'Part II：逐行对比详情（186条记录）', bold=True, fc="FFFFFF", fill=S_FILL, size=11, align='left')
ws2.row_dimensions[1].height = 22

# 列顺序（严格按要求）：
# SampleSeq | Activity(完整) |
# human_ann | human_fin | human_tp |
# deep_ann | deep_fin | deep_tp | deep_ann_TP/FP/FN/TN | deep_fin_TP/FP/FN/TN | deep_tp_TP/FP/FN/TN |
# mini_ann | mini_fin | mini_tp | mini_ann_TP/FP/FN/TN | mini_fin_TP/FP/FN/TN | mini_tp_TP/FP/FN/TN |
# mimo_ann | mimo_fin | mimo_tp | mimo_ann_TP/FP/FN/TN | mimo_fin_TP/FP/FN/TN | mimo_tp_TP/FP/FN/TN

col_headers = [
    'SampleSeq', 'Activity',
    'human_ann', 'human_fin', 'human_tp',
    # DeepSeek
    'deep_ann', 'deep_fin', 'deep_tp',
    'deep_ann_TP', 'deep_ann_FP', 'deep_ann_FN', 'deep_ann_TN',
    'deep_fin_TP', 'deep_fin_FP', 'deep_fin_FN', 'deep_fin_TN',
    'deep_tp_TP', 'deep_tp_FP', 'deep_tp_FN', 'deep_tp_TN',
    # MiniMax
    'mini_ann', 'mini_fin', 'mini_tp',
    'mini_ann_TP', 'mini_ann_FP', 'mini_ann_FN', 'mini_ann_TN',
    'mini_fin_TP', 'mini_fin_FP', 'mini_fin_FN', 'mini_fin_TN',
    'mini_tp_TP', 'mini_tp_FP', 'mini_tp_FN', 'mini_tp_TN',
    # Mimo
    'mimo_ann', 'mimo_fin', 'mimo_tp',
    'mimo_ann_TP', 'mimo_ann_FP', 'mimo_ann_FN', 'mimo_ann_TN',
    'mimo_fin_TP', 'mimo_fin_FP', 'mimo_fin_FN', 'mimo_fin_TN',
    'mimo_tp_TP', 'mimo_tp_FP', 'mimo_tp_FN', 'mimo_tp_TN',
]

hdr_row = 2
for c, h in enumerate(col_headers, 1):
    cell = ws2.cell(row=hdr_row, column=c, value=h)
    cell.fill = G_FILL
    cell.font = Font(name='Courier New', size=8, bold=True, color="1A1A18")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN
ws2.row_dimensions[hdr_row].height = 40

# 数据行
data_start = 3
for i in range(len(df)):
    row = df.iloc[i]
    r = data_start + i
    rf = W_FILL if i % 2 == 0 else ALT_FILL

    row_vals = [
        # identifiers
        row['SampleSeq'], row['Activity'],
        # human labels
        to_val(row['human_ann']), to_val(row['human_fin']), to_val(row['human_tp']),
        # DeepSeek predictions
        to_val(row['deep_ann']), to_val(row['deep_fin']), to_val(row['deep_tp']),
        # DeepSeek confusion matrices - ann_related
        to_val(row['deep_ann_tp']), to_val(row['deep_ann_fp']), to_val(row['deep_ann_fn']), to_val(row['deep_ann_tn']),
        # DeepSeek confusion matrices - ann_fin_flag
        to_val(row['deep_fin_tp']), to_val(row['deep_fin_fp']), to_val(row['deep_fin_fn']), to_val(row['deep_fin_tn']),
        # DeepSeek confusion matrices - third_party_flag
        to_val(row['deep_tp_tp']), to_val(row['deep_tp_fp']), to_val(row['deep_tp_fn']), to_val(row['deep_tp_tn']),
        # MiniMax predictions
        to_val(row['mini_ann']), to_val(row['mini_fin']), to_val(row['mini_tp']),
        # MiniMax confusion matrices - ann_related
        to_val(row['mini_ann_tp']), to_val(row['mini_ann_fp']), to_val(row['mini_ann_fn']), to_val(row['mini_ann_tn']),
        # MiniMax confusion matrices - ann_fin_flag
        to_val(row['mini_fin_tp']), to_val(row['mini_fin_fp']), to_val(row['mini_fin_fn']), to_val(row['mini_fin_tn']),
        # MiniMax confusion matrices - third_party_flag
        to_val(row['mini_tp_tp']), to_val(row['mini_tp_fp']), to_val(row['mini_tp_fn']), to_val(row['mini_tp_tn']),
        # Mimo predictions
        to_val(row['mimo_ann']), to_val(row['mimo_fin']), to_val(row['mimo_tp']),
        # Mimo confusion matrices - ann_related
        to_val(row['mimo_ann_tp']), to_val(row['mimo_ann_fp']), to_val(row['mimo_ann_fn']), to_val(row['mimo_ann_tn']),
        # Mimo confusion matrices - ann_fin_flag
        to_val(row['mimo_fin_tp']), to_val(row['mimo_fin_fp']), to_val(row['mimo_fin_fn']), to_val(row['mimo_fin_tn']),
        # Mimo confusion matrices - third_party_flag
        to_val(row['mimo_tp_tp']), to_val(row['mimo_tp_fp']), to_val(row['mimo_tp_fn']), to_val(row['mimo_tp_tn']),
    ]

    for c, v in enumerate(row_vals, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.fill = rf
        cell.font = Font(name='Courier New', size=9, color="1A1A18")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN

    ws2.row_dimensions[r].height = 18

# 末行汇总 F1
last_r = data_start + len(df)

# 第 last_r 行：写入 F1 表头（不合并，直接从 A 列开始）
f1_summary_hdr = ['F1 Score汇总', '模型', 'ann_related F1', 'ann_fin_flag F1', 'third_party_flag F1', '平均F1']
for ci, h in enumerate(f1_summary_hdr, 1):
    cell = ws2.cell(row=last_r, column=ci, value=h)
    cell.fill = H_FILL
    cell.font = Font(name='Courier New', size=9, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN

# last_r+1 到 last_r+3：三个模型的汇总数据
for ri, (mk, rv) in enumerate(results.items(), last_r + 1):
    cells_data = ['', rv['name'],
                  'N/A' if rv['f1_ann'] is None else f"{rv['f1_ann']:.4f}",
                  'N/A' if rv['f1_fin'] is None else f"{rv['f1_fin']:.4f}",
                  'N/A' if rv['f1_tp'] is None else f"{rv['f1_tp']:.4f}",
                  'N/A' if rv['avg'] is None else f"{rv['avg']:.4f}"]
    for ci, v in enumerate(cells_data, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell.fill = GR_FILL if ci == 6 else W_FILL
        cell.font = Font(name='Courier New', size=9, bold=(ci == 6), color="3B6D11" if ci == 6 else "1A1A18")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN

for ri in range(last_r, last_r + 4):
    ws2.row_dimensions[ri].height = 20

# 设置列宽
col_widths = [10, 50] + [8]*3 + [8]*3 + [8]*12 + [8]*3 + [8]*12 + [8]*3 + [8]*12
for c, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# 保存
output = 'validation_夏思远_2023110537_G02.xlsx'
wb.save(output)
print(f"\n已保存: {output}")
print(f"Sheet1: 模型对比总览")
print(f"Sheet2: 详细逐条记录 {len(df)}行 + F1汇总")