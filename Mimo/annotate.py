# -*- coding: utf-8 -*-
"""
AI辅助分类财务舞弊 - 自动标注脚本
基于作业说明_update.html中的规则对blank_G02.xlsx进行标注
"""

import sys
import io
import re
import json
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ============================================================
# 第一步：判断是否年报相关
# ============================================================

def is_annual_report_related(activity, violation_type):
    """
    判断违规行为是否导致年度报告中的信息存在错误或遗漏。
    返回: (bool, str) - (是否年报相关, 判断理由)
    """
    if not activity:
        return False, "无Activity文本"

    text = activity.lower()

    # ---- 排除规则（优先判断） ----

    # 1. 内幕交易 - 通常不影响年报
    if violation_type and '内幕交易' in violation_type:
        # 检查是否有明确的年报虚假记载
        has_annual_report_issue = False
        # 检查是否同时有其他年报相关的违规类型
        if violation_type and any(kw in violation_type for kw in ['虚假记载', '虚构利润', '欺诈上市', '重大遗漏']):
            has_annual_report_issue = True
        if not has_annual_report_issue:
            return False, "内幕交易类违规，不影响年报"

    # 2. 违规买卖股票（短线交易、窗口期交易等）- 通常不影响年报
    if violation_type and '违规买卖股票' in violation_type:
        # 检查是否有明确的年报相关违规
        has_annual_report_issue = False
        if violation_type and any(kw in violation_type for kw in ['虚假记载', '虚构利润', '欺诈上市']):
            has_annual_report_issue = True
        # 检查文本中是否有年报相关描述
        if '年度报告' in activity or '年报' in activity:
            # 检查年报是否有问题
            if any(kw in activity for kw in ['虚假记载', '虚增', '虚减', '少计', '多计', '不准确', '不真实', '遗漏']):
                has_annual_report_issue = True
        if not has_annual_report_issue:
            return False, "违规买卖股票，不影响年报"

    # 3. 仅涉及半年报/季报，不涉及年报
    if ('半年度报告' in activity or '半年报' in activity or '季度报告' in activity or '季报' in activity):
        if '年度报告' not in activity and '年报' not in activity:
            # 检查是否有间接影响年报的描述
            if not any(kw in activity for kw in ['年度报告', '年报']):
                return False, "仅涉及半年报/季报，未涉及年报"

    # 4. 未缴或少缴税款 - 通常不影响年报
    if violation_type and '未缴或少缴税款' in violation_type:
        # 检查是否有明确的年报相关描述
        if '年度报告' not in activity and '年报' not in activity:
            return False, "未缴税款类违规，未涉及年报"

    # 5. 安全生产事故 - 通常不影响年报
    if '安全事故' in activity and '安全生产' in activity:
        if '年度报告' not in activity and '年报' not in activity:
            return False, "安全生产事故，未涉及年报"

    # 6. 化妆品/药品/产品质量违规 - 通常不影响年报
    if any(kw in activity for kw in ['化妆品', '药品销售', '生产工艺', '配方投料']):
        if '年度报告' not in activity and '年报' not in activity:
            if not any(kw in violation_type for kw in ['虚假记载', '虚构利润', '重大遗漏']):
                return False, "产品质量违规，未涉及年报"

    # 7. 董事未出席会议 - 通常不影响年报
    if '未出席' in activity and '董事会' in activity:
        if '年度报告' not in activity and '年报' not in activity:
            return False, "董事未出席，不影响年报"

    # 8. 增持承诺未履行 - 通常不影响年报
    if '增持' in activity and '承诺' in activity:
        if '年度报告' not in activity and '年报' not in activity:
            if not any(kw in violation_type for kw in ['虚假记载', '虚构利润']):
                return False, "增持承诺未履行，不影响年报"

    # 9. 工程违法转包 - 通常不影响年报
    if '违法转包' in activity:
        return False, "工程违法转包，不影响年报"

    # ---- 纳入规则（判断是否年报相关） ----

    # 规则1：文本中明确提到年度报告/年报有问题
    annual_report_keywords = ['年度报告', '年报']
    has_annual_report_mention = any(kw in activity for kw in annual_report_keywords)

    if has_annual_report_mention:
        # 检查年报是否有问题
        issue_indicators = [
            '虚假记载', '虚增', '虚减', '少计', '多计', '不准确', '不真实',
            '重大遗漏', '误导性陈述', '虚假', '不完整', '错误',
            '未披露', '未在.*披露', '未按规定披露',
            '少计提', '多计提', '少确认', '多确认',
            '提前确认', '推迟确认', '不当',
            '虚构', '编造', '隐瞒',
        ]
        for indicator in issue_indicators:
            if re.search(indicator, activity):
                return True, f"年度报告中存在{indicator}问题"

        # 检查是否有业绩预告与年报差异
        if '业绩预告' in activity and ('年度报告' in activity or '年报' in activity):
            if any(kw in activity for kw in ['差异', '修正', '更正', '不准确']):
                return True, "业绩预告与年报存在差异"

    # 规则2：违规类型直接暗示年报问题
    violation_type_keywords = ['虚假记载', '虚构利润', '欺诈上市', '一般会计处理不当']
    for kw in violation_type_keywords:
        if violation_type and kw in violation_type:
            # 进一步检查是否涉及年报
            if any(akw in activity for akw in ['年度报告', '年报', '年度', '财务报表']):
                return True, f"违规类型'{kw}'涉及年报"
            # 对于虚构利润/欺诈上市，通常涉及年报
            if kw in ['虚构利润', '欺诈上市']:
                return True, f"违规类型'{kw}'涉及年报"

    # 规则3：披露不实涉及财务信息
    if violation_type and '披露不实' in violation_type:
        if any(kw in activity for kw in ['年度报告', '年报', '财务报表', '半年度报告']):
            return True, "披露不实涉及定期报告"

    # 规则4：业绩预告/业绩快报问题（通常影响年报）
    if '业绩预告' in activity or '业绩快报' in activity:
        if '年度报告' in activity or '年报' in activity:
            return True, "业绩预告/快报与年报相关"
        # 业绩预告通常与年报相关
        if any(kw in activity for kw in ['净利润', '营业收入', '归属于上市公司股东']):
            # 检查是否提到年度
            year_pattern = r'(20\d{2})年度|20\d{2}年'
            years = re.findall(year_pattern, activity)
            if years:
                return True, "业绩预告涉及年度业绩"

    # 规则5：涉及年度财务数据错误
    if any(kw in activity for kw in ['年度报告', '年报']):
        if any(kw in activity for kw in ['虚增', '虚减', '少计', '多计', '少确认', '多确认']):
            return True, "年度财务数据存在虚假记载"

    # 规则6：招股说明书虚假记载（欺诈上市）
    if '招股说明书' in activity and any(kw in activity for kw in ['虚假记载', '虚增', '虚减', '编造', '隐瞒']):
        return True, "招股说明书存在虚假记载"

    # 规则7：涉及年报的关联交易/关联方问题
    if ('年度报告' in activity or '年报' in activity):
        if any(kw in activity for kw in ['关联交易', '关联方', '关联关系']):
            if any(kw in activity for kw in ['未披露', '不准确', '不完整', '重大遗漏']):
                return True, "年报中关联交易披露问题"

    # 规则8：募集资金问题涉及年报
    if '募集资金' in activity and ('年度报告' in activity or '年报' in activity):
        if any(kw in activity for kw in ['未披露', '不准确', '不完整']):
            return True, "年报中募集资金披露问题"

    return False, "未发现年报相关问题"


def extract_years(activity):
    """从Activity文本中提取受影响的年份。"""
    years = set()

    # 匹配模式：2017年、2018年年度报告
    patterns = [
        r'(20\d{2})年年度报告',
        r'(20\d{2})年年报',
        r'(20\d{2})年度',
        r'20\d{2}年至(20\d{2})年',
        r'(20\d{2})年',
    ]

    for pattern in patterns:
        matches = re.findall(pattern, activity)
        for match in matches:
            if isinstance(match, tuple):
                for m in match:
                    if m and m.startswith('20'):
                        years.add(int(m))
            elif match and match.startswith('20'):
                years.add(int(match))

    # 特殊处理范围：2016年至2019年
    range_patterns = [
        r'(20\d{2})年至(20\d{2})年',
        r'(20\d{2})年-(20\d{2})年',
    ]
    for pattern in range_patterns:
        matches = re.findall(pattern, activity)
        for match in matches:
            start_year = int(match[0])
            end_year = int(match[1])
            for y in range(start_year, end_year + 1):
                years.add(y)

    # 过滤：只保留2000年之后的年份，且不超过当前年份
    years = {y for y in years if 2000 <= y <= 2026}

    return sorted(list(years)) if years else None


# ============================================================
# 第二步：判断是否影响财务信息
# ============================================================

def is_financial_info(activity, violation_type):
    """
    判断年报中受影响的是财务信息还是非财务信息。
    财务信息指财务报表及相关附注中的科目金额。
    """
    if not activity:
        return False, "无Activity文本"

    # 财务相关的违规类型
    financial_violation_types = [
        '虚假记载', '虚构利润', '一般会计处理不当',
        '欺诈上市', '披露不实'
    ]

    for fvt in financial_violation_types:
        if violation_type and fvt in violation_type:
            # 进一步确认是否涉及财务数据
            if any(kw in activity for kw in [
                '营业收入', '净利润', '利润', '收入', '成本', '费用',
                '资产', '负债', '权益', '应收账款', '存货', '固定资产',
                '虚增', '虚减', '少计', '多计', '少确认', '多确认',
                '提前确认', '推迟确认', '坏账', '减值', '折旧',
                '财务报表', '财务报告', '财务数据',
                '业绩预告', '业绩快报', '年报'
            ]):
                return True, f"违规类型'{fvt}'涉及财务数据"

    # 关键词匹配 - 财务造假相关
    financial_keywords = [
        '虚增营业收入', '虚增收入', '虚增利润', '虚减利润',
        '虚增净利润', '虚减净利润', '虚增资产', '虚减资产',
        '少计提', '多计提', '少确认', '多确认',
        '提前确认收入', '推迟确认收入', '提前确认', '推迟确认',
        '虚增应收账款', '虚增存货', '虚增固定资产',
        '少计费用', '多计费用', '少计成本', '多计成本',
        '虚假记载', '虚假合同', '虚构交易', '虚构业务',
        '伪造', '编造', '不实', '不准确',
        '收入确认', '费用冲减', '成本核算',
        '坏账准备', '资产减值', '商誉减值',
        '折旧', '摊销',
        '关联交易.*金额', '资金占用',
    ]

    for kw in financial_keywords:
        if re.search(kw, activity):
            return True, f"涉及财务关键词: {kw}"

    # 非财务信息的关键词（排除）
    non_financial_keywords = [
        '公司治理', '内部控制评价报告', '股东大会',
        '董事会决议', '信息披露义务', '审议程序',
        '减持计划', '增持承诺', '权益变动',
    ]

    has_non_financial = any(kw in activity for kw in non_financial_keywords)
    has_financial = any(kw in activity for kw in [
        '营业收入', '净利润', '利润总额', '资产', '负债',
        '应收账款', '存货', '成本', '费用'
    ])

    if has_non_financial and not has_financial:
        return False, "主要涉及非财务信息"

    return False, "未发现明确的财务信息影响"


# ============================================================
# 第三步：识别受影响的会计要素
# ============================================================

def identify_financial_elements(activity):
    """
    识别受影响的会计要素（六大要素）。
    返回: (list, list) - (年份-要素列表, 元素列表)
    """
    if not activity:
        return None, []

    # 定义关键词到会计要素的映射
    element_mapping = {
        '资产': [
            '应收账款', '应收票据', '其他应收款', '预付账款',
            '存货', '固定资产', '无形资产', '在建工程', '商誉',
            '长期投资', '股权投资', '投资性房地产', '开发支出',
            '货币资金', '银行存款', '资金', '资金占用',
            '坏账', '减值准备', '资产减值',
            '借款', '贷款', '应收',
            '委托理财', '担保', '质押',
            '往来', '拆借', '保证金',
        ],
        '负债': [
            '应付账款', '应付票据', '其他应付款', '预收账款',
            '短期借款', '长期借款', '应付债券', '借款',
            '贷款', '担保', '债务',
            '预计负债', '递延收益',
            '负债',
        ],
        '所有者权益': [
            '实收资本', '资本公积', '盈余公积', '未分配利润',
            '所有者权益', '净资产', '股东权益',
            '配股', '送股', '转增',
        ],
        '收入': [
            '营业收入', '主营业务收入', '其他业务收入',
            '收入', '销售收入', '确认收入',
            '投资收益', '利息收入',
            '提前确认', '推迟确认',
        ],
        '费用': [
            '营业成本', '主营业务成本', '管理费用', '销售费用',
            '财务费用', '研发费用', '所得税费用',
            '成本', '费用', '折旧', '摊销',
            '坏账准备', '减值准备', '减值损失',
            '信用减值', '资产减值',
            '期间费用', '税金',
        ],
        '利润': [
            '净利润', '利润总额', '营业利润', '利润',
            '归属于上市公司股东的净利润',
            '扣除非经常性损益后的净利润', '扣非净利润',
            '综合收益',
        ],
    }

    found_elements = set()
    year_elements = {}

    # 首先尝试按年份提取要素
    year_pattern = r'(20\d{2})年'
    year_matches = list(re.finditer(year_pattern, activity))

    for match in year_matches:
        year = int(match.group(1))
        if year < 2000 or year > 2026:
            continue

        # 获取该年份之后的文本片段
        start_pos = match.start()
        # 查找下一个年份或文本结束
        next_year_match = re.search(r'(20\d{2})年', activity[match.end():])
        if next_year_match:
            end_pos = match.end() + next_year_match.start()
        else:
            end_pos = len(activity)

        segment = activity[start_pos:end_pos]

        # 检查该段文本中涉及的会计要素
        for element, keywords in element_mapping.items():
            for kw in keywords:
                if kw in segment:
                    if year not in year_elements:
                        year_elements[year] = set()
                    year_elements[year].add(element)
                    found_elements.add(element)
                    break

    # 如果没有按年份找到要素，则全局搜索
    if not found_elements:
        for element, keywords in element_mapping.items():
            for kw in keywords:
                if kw in activity:
                    found_elements.add(element)
                    break

    # 格式化输出
    if year_elements:
        ann_fin_info = []
        for year in sorted(year_elements.keys()):
            elements = sorted(list(year_elements[year]))
            ann_fin_info.append({"year": year, "elements": elements})
        return ann_fin_info, list(found_elements)
    elif found_elements:
        return None, list(found_elements)
    else:
        return None, []


# ============================================================
# 第四步：识别第三方配合造假
# ============================================================

def identify_third_party(activity, violation_type):
    """
    识别是否存在第三方配合财务舞弊。
    返回: (bool, list) - (是否有第三方配合, 第三方列表)
    """
    if not activity:
        return False, None

    third_parties = []

    # 第三方配合的关键词
    collusion_keywords = [
        '配合签订虚假', '配合开具', '配合.*虚假',
        '签订虚假合同', '签订虚假采购合同', '签订虚假销售合同',
        '虚开.*发票', '虚开增值税', '虚开销售发票',
        '虚假.*验收', '不实.*验收', '虚假.*证明',
        '伪造.*凭证', '伪造.*合同', '伪造.*发票',
        '资金回流', '资金循环', '回流',
        '出借账户', '代持股份', '代持',
        '配合.*造假', '配合.*舞弊', '配合.*虚增',
        '协助.*造假', '协助.*舞弊',
        '配合虚构', '配合虚增', '协助虚构',
        '串通', '合谋',
    ]

    has_collusion = False
    for kw in collusion_keywords:
        if re.search(kw, activity):
            has_collusion = True
            break

    if not has_collusion:
        return False, None

    # 提取第三方名称和类型
    # 模式1：供应商/客户/银行 + 公司名称 + 动作
    party_patterns = [
        # 供应商XXX配合签订虚假合同
        r'(?:供应商|供方|卖方)([^\x00-\x7F]{2,30}(?:有限公司|股份有限公司|集团|公司))',
        # 客户XXX配合
        r'(?:客户|买方|购方)([^\x00-\x7F]{2,30}(?:有限公司|股份有限公司|集团|公司))',
        # XXX有限公司配合签订
        r'([^\x00-\x7F]{2,30}(?:有限公司|股份有限公司))配合',
        # XXX公司提供虚假
        r'([^\x00-\x7F]{2,30}(?:有限公司|股份有限公司|公司))提供虚假',
        # 与XXX签订虚假
        r'与([^\x00-\x7F]{2,30}(?:有限公司|股份有限公司|公司))签订虚假',
        # 自然人XXX代持
        r'自然人([^\x00-\x7F]{2,6})代持',
        # XXX会计师事务所出具虚假
        r'([^\x00-\x7F]{2,30}(?:会计师事务所|评估机构))出具虚假',
    ]

    for pattern in party_patterns:
        matches = re.findall(pattern, activity)
        for match in matches:
            name = match.strip()
            if name and len(name) >= 2:
                # 判断类型
                ptype = "其他企业"
                if '供应商' in activity[max(0, activity.find(name)-20):activity.find(name)+len(name)+20]:
                    ptype = "供应商"
                elif '客户' in activity[max(0, activity.find(name)-20):activity.find(name)+len(name)+20]:
                    ptype = "客户"
                elif '银行' in activity[max(0, activity.find(name)-20):activity.find(name)+len(name)+20]:
                    ptype = "银行/金融机构"
                elif '会计师' in name:
                    ptype = "会计师事务所"
                elif '评估' in name:
                    ptype = "评估机构"
                elif '券商' in name or '证券' in name:
                    ptype = "券商/保荐机构"
                elif '自然人' in activity[max(0, activity.find(name)-10):activity.find(name)]:
                    ptype = "自然人"

                # 获取角色描述
                role_start = activity.find(name)
                role_end = min(role_start + 100, len(activity))
                role_text = activity[role_start:role_end]
                # 截取到下一个句号或逗号
                for sep in ['。', '；', '，', ',', ';']:
                    sep_pos = role_text.find(sep)
                    if sep_pos > 0:
                        role_text = role_text[:sep_pos]
                        break

                # 简化角色描述
                role = "配合财务舞弊"
                if '虚假合同' in role_text:
                    role = "签订虚假合同"
                elif '虚假验收' in role_text:
                    role = "配合开具虚假验收证明"
                elif '资金回流' in role_text or '回流' in role_text:
                    role = "配合资金回流"
                elif '虚开' in role_text:
                    role = "虚开票据"
                elif '代持' in role_text:
                    role = "代持股份"

                if name not in [p.get('name') for p in third_parties]:
                    third_parties.append({
                        "name": name,
                        "type": ptype,
                        "role": role
                    })

    if third_parties:
        return True, third_parties

    # 如果检测到配合关键词但未提取到具体名称，标记为有第三方但名称不详
    if has_collusion:
        # 再次尝试更宽泛的名称提取
        company_pattern = r'([一-龥]{2,20}(?:有限公司|股份有限公司|集团|公司))'
        companies = re.findall(company_pattern, activity)
        # 过滤掉上市公司自身
        filtered = []
        for c in companies:
            # 排除常见上市公司自身名称模式
            if '本公司' not in c and '公司' in c:
                filtered.append(c)

        if filtered:
            # 取第一个作为第三方
            name = filtered[0]
            ptype = "其他企业"
            if '供应商' in activity:
                ptype = "供应商"
            elif '客户' in activity:
                ptype = "客户"

            third_parties.append({
                "name": name,
                "type": ptype,
                "role": "配合财务舞弊（文本中未出现完整名称）"
            })
            return True, third_parties

    return False, None


# ============================================================
# 主处理函数
# ============================================================

def process_row(activity, violation_type):
    """处理单行数据，返回六个字段的值。"""
    # 默认值
    ann_related = 0
    ann_year = None
    ann_fin_flag = None
    ann_fin_info = None
    third_party_flag = None
    third_party_list = None

    # Step 1: 判断年报相关性
    is_related, reason = is_annual_report_related(activity, violation_type)
    ann_related = 1 if is_related else 0

    if ann_related == 0:
        return ann_related, ann_year, ann_fin_flag, ann_fin_info, third_party_flag, third_party_list

    # Step 2: 提取受影响年份
    ann_year = extract_years(activity)

    # Step 3: 判断是否影响财务信息
    is_financial, fin_reason = is_financial_info(activity, violation_type)
    ann_fin_flag = 1 if is_financial else 0

    if ann_fin_flag == 0:
        return ann_related, ann_year, ann_fin_flag, ann_fin_info, third_party_flag, third_party_list

    # Step 4: 识别会计要素
    ann_fin_info, elements = identify_financial_elements(activity)

    # 如果ann_fin_info为空但有elements，按年份组织
    if ann_fin_info is None and elements and ann_year:
        ann_fin_info = []
        for year in ann_year:
            ann_fin_info.append({"year": year, "elements": elements})

    # Step 5: 判断第三方配合
    has_third_party, third_party_list = identify_third_party(activity, violation_type)
    third_party_flag = 1 if has_third_party else 0

    return ann_related, ann_year, ann_fin_flag, ann_fin_info, third_party_flag, third_party_list


# ============================================================
# 主程序
# ============================================================

def main():
    input_file = 'blank_G02.xlsx'
    output_file = 'Mimo/mimo_ver.xlsx'

    print("正在读取数据...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in ws[1]]
    print(f"表头: {headers}")

    # 找到Activity列的索引
    activity_col = None
    violation_type_col = None
    for i, h in enumerate(headers):
        if h == 'Activity':
            activity_col = i + 1  # openpyxl是1-based
        if h == 'ViolationType':
            violation_type_col = i + 1

    if activity_col is None:
        print("错误：未找到Activity列")
        return

    print(f"Activity列: 第{activity_col}列")
    if violation_type_col:
        print(f"ViolationType列: 第{violation_type_col}列")

    # 处理每一行
    results = []
    total_rows = ws.max_row - 1

    for row_idx in range(2, ws.max_row + 1):
        activity = ws.cell(row=row_idx, column=activity_col).value
        violation_type = ws.cell(row=row_idx, column=violation_type_col).value if violation_type_col else ""

        if not activity:
            activity = ""

        if not violation_type:
            violation_type = ""

        # 处理行
        ann_related, ann_year, ann_fin_flag, ann_fin_info, third_party_flag, third_party_list = process_row(activity, violation_type)

        # 格式化输出
        ann_year_str = json.dumps(ann_year, ensure_ascii=False) if ann_year else "null"
        ann_fin_info_str = json.dumps(ann_fin_info, ensure_ascii=False) if ann_fin_info else "null"
        third_party_list_str = json.dumps(third_party_list, ensure_ascii=False) if third_party_list else "null"

        results.append({
            'row': row_idx,
            'ann_related': ann_related,
            'ann_year': ann_year_str,
            'ann_fin_flag': ann_fin_flag,
            'ann_fin_info': ann_fin_info_str,
            'third_party_flag': third_party_flag,
            'third_party_list': third_party_list_str,
        })

        # 进度显示
        if (row_idx - 1) % 20 == 0 or row_idx == ws.max_row:
            print(f"  已处理 {row_idx - 1}/{total_rows} 行...")

    # 输出统计
    print("\n=== 标注统计 ===")
    ann_related_count = sum(1 for r in results if r['ann_related'] == 1)
    ann_fin_count = sum(1 for r in results if r['ann_fin_flag'] == 1)
    third_party_count = sum(1 for r in results if r['third_party_flag'] == 1)
    print(f"总行数: {total_rows}")
    print(f"ann_related=1: {ann_related_count}")
    print(f"ann_fin_flag=1: {ann_fin_count}")
    print(f"third_party_flag=1: {third_party_count}")

    # 输出到新Excel文件
    print(f"\n正在生成输出文件: {output_file}")
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'sheet1'

    # 写入表头
    out_headers = [
        'Activity', 'ann_related', 'ann_year', 'ann_fin_flag',
        'ann_fin_info', 'third_party_flag', 'third_party_list'
    ]
    for col, header in enumerate(out_headers, 1):
        out_ws.cell(row=1, column=col, value=header)

    # 写入数据
    for row_idx, result in enumerate(results, 2):
        # 获取原始Activity
        activity = ws.cell(row=row_idx, column=activity_col).value or ""

        out_ws.cell(row=row_idx, column=1, value=activity)
        out_ws.cell(row=row_idx, column=2, value=result['ann_related'])
        out_ws.cell(row=row_idx, column=3, value=result['ann_year'])
        out_ws.cell(row=row_idx, column=4, value=result['ann_fin_flag'])
        out_ws.cell(row=row_idx, column=5, value=result['ann_fin_info'])
        out_ws.cell(row=row_idx, column=6, value=result['third_party_flag'])
        out_ws.cell(row=row_idx, column=7, value=result['third_party_list'])

    # 调整列宽
    out_ws.column_dimensions['A'].width = 80
    out_ws.column_dimensions['B'].width = 12
    out_ws.column_dimensions['C'].width = 18
    out_ws.column_dimensions['D'].width = 14
    out_ws.column_dimensions['E'].width = 60
    out_ws.column_dimensions['F'].width = 16
    out_ws.column_dimensions['G'].width = 60

    out_wb.save(output_file)
    print(f"\n输出文件已保存: {output_file}")

    # 输出详细结果（前20行）
    print("\n=== 前20行标注结果 ===")
    for result in results[:20]:
        print(f"Row {result['row']}: ann_related={result['ann_related']}, "
              f"ann_year={result['ann_year']}, "
              f"ann_fin_flag={result['ann_fin_flag']}, "
              f"third_party_flag={result['third_party_flag']}")

    # 输出ann_related=1的行
    print("\n=== ann_related=1 的行 ===")
    for result in results:
        if result['ann_related'] == 1:
            print(f"Row {result['row']}: ann_year={result['ann_year']}, "
                  f"ann_fin_flag={result['ann_fin_flag']}, "
                  f"ann_fin_info={result['ann_fin_info'][:100] if result['ann_fin_info'] != 'null' else 'null'}, "
                  f"third_party_flag={result['third_party_flag']}")


if __name__ == '__main__':
    main()
