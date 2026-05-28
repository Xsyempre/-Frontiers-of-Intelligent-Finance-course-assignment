# -*- coding: utf-8 -*-
"""
对v2标注结果进行人工审查修正
"""

import sys
import io
import json
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 修正映射表：row -> (ann_related, ann_year, ann_fin_flag, ann_fin_info, third_party_flag, third_party_list)
# None 表示不修改该字段

corrections = {
    # Row 2: 年报问题仅涉及2018和2019
    2: (1, [2018, 2019], 1, [{"year": 2019, "elements": ["资产"]}], 0, None),

    # Row 8: 未披露关联交易涉及2019-2020年报
    8: (1, [2019, 2020], 1, [{"year": 2019, "elements": ["资产"]}, {"year": 2020, "elements": ["资产"]}], 0, None),

    # Row 10: 海航集团2015-2018年报未披露担保，2019-2020关联交易未披露
    10: (1, [2015, 2016, 2017, 2018, 2019, 2020], 1, [
        {"year": 2019, "elements": ["资产"]},
        {"year": 2020, "elements": ["资产"]}
    ], 0, None),

    # Row 14: 招股说明书+年报，欺诈上市期间2016-2019
    14: (1, [2016, 2017, 2018, 2019], 1, [
        {"year": 2016, "elements": ["收入", "利润"]},
        {"year": 2017, "elements": ["收入", "利润", "资产"]},
        {"year": 2018, "elements": ["收入", "利润", "资产"]},
        {"year": 2019, "elements": ["收入", "利润", "资产"]}
    ], 0, None),

    # Row 22: 报表合并不完整+资金占用，ann_year应为null（未明确年份）
    22: (1, None, 1, None, 0, None),

    # Row 27: 2015-2016年虚增，2018年是更正年份
    27: (1, [2015, 2016], 1, [
        {"year": 2015, "elements": ["收入", "利润"]},
        {"year": 2016, "elements": ["收入", "利润"]}
    ], 0, None),

    # Row 31: 募集资金+内控问题，无明确年份
    31: (1, None, 1, None, 0, None),

    # Row 33: 未披露关联交易
    33: (1, [2022, 2023], 1, [{"year": 2022, "elements": ["资产"]}, {"year": 2023, "elements": ["资产"]}], 0, None),

    # Row 37: 2004年年报未及时披露亏损
    37: (1, [2004], 0, None, None, None),

    # Row 52: 关联交易未及时披露+业绩预告违规
    52: (1, [2010], 0, None, None, None),

    # Row 53: 2017年业绩预告修正不及时
    53: (1, [2017], 1, [{"year": 2017, "elements": ["利润"]}], 0, None),

    # Row 75: 未披露关联关系，涉及2014年年报
    75: (1, [2014], 0, None, None, None),

    # Row 108: 康美药业主要是2019年业绩预告问题
    108: (1, [2019], 1, [{"year": 2019, "elements": ["利润"]}], 0, None),

    # Row 135: 移除2014（是文件号不是年份）
    135: (1, [2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023], 1, [
        {"year": 2016, "elements": ["资产"]},
        {"year": 2017, "elements": ["资产"]},
        {"year": 2019, "elements": ["资产"]},
        {"year": 2021, "elements": ["利润", "资产"]},
        {"year": 2022, "elements": ["负债"]},
        {"year": 2023, "elements": ["资产"]}
    ], 0, None),

    # Row 142: 关联交易未披露涉及2020-2021年报
    142: (1, [2020, 2021], 0, None, None, None),

    # Row 155: 修正第三方（金荔科技自身不是第三方）
    155: (1, [2002, 2003, 2004], 1, [
        {"year": 2002, "elements": ["收入", "利润"]},
        {"year": 2003, "elements": ["收入", "利润"]},
        {"year": 2004, "elements": ["收入", "利润"]}
    ], 0, None),

    # Row 156: 瑞斯康达专网通信，第三方是隋某力控制的企业
    156: (1, [2019, 2020], 1, [
        {"year": 2019, "elements": ["收入", "利润", "费用"]},
        {"year": 2020, "elements": ["收入", "利润", "费用"]}
    ], 1, [{"name": "隋某力控制的企业", "type": "其他企业", "role": "专网通信虚假自循环业务上下游配合"}]),

    # Row 167: 资产冻结未及时披露，涉及2017年报
    167: (1, [2017], 0, None, None, None),

    # Row 173: 2020年年报会计差错更正
    173: (1, [2020], 1, [{"year": 2020, "elements": ["收入", "费用"]}], 0, None),

    # Row 178: 2019年半年报问题，但更正影响2019年年度数据
    178: (1, [2019], 1, [{"year": 2019, "elements": ["利润"]}], 0, None),

    # 额外修正：一些漏判的行

    # Row 3: 合同管理+内控问题，2021年报审计意见有问题
    # 实际上是内控问题，不是年报数据错误，保持ann_related=0

    # Row 11: 资产评估问题，商誉减值
    # 不直接涉及年报，保持ann_related=0

    # Row 35: 债务重组公告问题
    # 不直接涉及年报，保持ann_related=0

    # Row 39: 未披露关联公司情况，可能涉及年报
    # 实际上是控股股东的问题，不直接涉及年报数据

    # Row 44: 独立财务顾问出具虚假记载
    # 不直接涉及年报

    # Row 54: 募集资金+半年报问题
    # 半年报不是年报，保持ann_related=0

    # Row 61: 公司治理+财务信息披露问题
    # 2010年度财务报告披露不完整
    61: (1, [2010], 1, [{"year": 2010, "elements": ["利润", "资产"]}], 0, None),

    # Row 64: 会计估计变更不及时，导致2019年多计利润
    64: (1, [2019], 1, [{"year": 2019, "elements": ["利润", "费用"]}], 0, None),

    # Row 73: 收购标的公司财务变化未及时披露，涉及2012年年报
    73: (1, [2012], 1, [{"year": 2012, "elements": ["利润", "资产"]}], 0, None),

    # Row 92: 慧辰股份招股说明书+年报虚假记载，子公司虚构业务
    # 第三方：信唐普华通过虚构与第三方业务
    92: (1, [2018, 2019, 2020, 2021, 2022], 1, [
        {"year": 2018, "elements": ["收入", "利润"]},
        {"year": 2019, "elements": ["收入", "利润"]},
        {"year": 2020, "elements": ["收入", "利润"]},
        {"year": 2021, "elements": ["收入", "利润"]},
        {"year": 2022, "elements": ["收入", "利润"]}
    ], 1, [{"name": "北京信唐普华科技有限公司", "type": "其他企业", "role": "通过签订无商业实质销售合同虚增收入和利润"}]),

    # Row 119: 闽越花雕2004年年报虚假记载
    119: (1, [2004], 1, [{"year": 2004, "elements": ["利润", "费用"]}], 0, None),

    # Row 128: 收入确认跨期，2014年年报
    128: (1, [2014], 1, [{"year": 2014, "elements": ["收入", "利润"]}], 0, None),

    # Row 146: 国民技术2018年业绩预告/快报与年报差异
    146: (1, [2018], 1, [{"year": 2018, "elements": ["利润", "费用"]}], 0, None),

    # Row 180: 年报中子公司披露不完整+关联交易不完整
    180: (1, [2011], 1, [{"year": 2011, "elements": ["资产"]}], 0, None),
}


def main():
    input_file = 'Mimo/mimo_ver.xlsx'
    output_file = 'Mimo/mimo_ver.xlsx'

    print("正在读取v2标注结果...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    print(f"总行数: {ws.max_row - 1}")

    # 应用修正
    corrected_count = 0
    for row_idx, correction in corrections.items():
        excel_row = row_idx  # corrections dict中的key就是Excel行号

        ann_related, ann_year, ann_fin_flag, ann_fin_info, third_party_flag, third_party_list = correction

        if ann_related is not None:
            ws.cell(row=excel_row, column=2, value=ann_related)

        if ann_year is not None:
            ws.cell(row=excel_row, column=3, value=json.dumps(ann_year, ensure_ascii=False))
        elif ann_year is None and correction[0] == 1:
            ws.cell(row=excel_row, column=3, value="null")

        if ann_fin_flag is not None:
            ws.cell(row=excel_row, column=4, value=ann_fin_flag)

        if ann_fin_info is not None:
            ws.cell(row=excel_row, column=5, value=json.dumps(ann_fin_info, ensure_ascii=False))
        elif ann_fin_info is None and correction[2] == 1:
            ws.cell(row=excel_row, column=5, value="null")

        if third_party_flag is not None:
            ws.cell(row=excel_row, column=6, value=third_party_flag)

        if third_party_list is not None:
            ws.cell(row=excel_row, column=7, value=json.dumps(third_party_list, ensure_ascii=False))
        elif third_party_list is None and correction[4] == 0:
            ws.cell(row=excel_row, column=7, value="null")

        corrected_count += 1

    print(f"已修正 {corrected_count} 行")

    # 统计修正后的结果
    ann_related_count = 0
    ann_fin_count = 0
    third_party_count = 0

    for row_idx in range(2, ws.max_row + 1):
        ar = ws.cell(row=row_idx, column=2).value
        af = ws.cell(row=row_idx, column=4).value
        tp = ws.cell(row=row_idx, column=6).value

        if ar == 1:
            ann_related_count += 1
        if af == 1:
            ann_fin_count += 1
        if tp == 1:
            third_party_count += 1

    print(f"\n=== 修正后统计 ===")
    print(f"ann_related=1: {ann_related_count}")
    print(f"ann_fin_flag=1: {ann_fin_count}")
    print(f"third_party_flag=1: {third_party_count}")

    wb.save(output_file)
    print(f"\n已保存: {output_file}")

    # 输出修正后的详细结果
    print("\n=== 修正后 ann_related=1 的行 ===")
    for row_idx in range(2, ws.max_row + 1):
        ar = ws.cell(row=row_idx, column=2).value
        if ar == 1:
            ay = ws.cell(row=row_idx, column=3).value
            af = ws.cell(row=row_idx, column=4).value
            ai = ws.cell(row=row_idx, column=5).value
            tp = ws.cell(row=row_idx, column=6).value
            tl = ws.cell(row=row_idx, column=7).value
            print(f"Row {row_idx-1}: ann_year={ay}, ann_fin_flag={af}, third_party_flag={tp}")


if __name__ == '__main__':
    main()
