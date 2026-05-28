# -*- coding: utf-8 -*-
"""
AI辅助分类财务舞弊 - 自动标注脚本 v2
改进版：更精确的年份提取、第三方检测和分类逻辑
"""

import sys
import io
import re
import json
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ============================================================
# 年份提取 - 更精确的版本
# ============================================================

def extract_years(activity):
    """
    从Activity文本中提取受影响的年份。
    只提取与年度报告/年报直接相关的年份。
    """
    if not activity:
        return None

    years = set()

    # 模式1：明确的"XXXX年年度报告"或"XXXX年年报"
    for m in re.finditer(r'(20\d{2})年(?:年度报告|年报)', activity):
        years.add(int(m.group(1)))

    # 模式2："XXXX年至XXXX年年度报告"
    for m in re.finditer(r'(20\d{2})年至(20\d{2})年(?:年度报告|年报)', activity):
        for y in range(int(m.group(1)), int(m.group(2)) + 1):
            years.add(y)

    # 模式3："《XXXX年年度报告》"
    for m in re.finditer(r'《(20\d{2})年(?:年度报告|年报)》', activity):
        years.add(int(m.group(1)))

    # 模式4："XXXX年度" + 财务/报告相关上下文
    for m in re.finditer(r'(20\d{2})年度', activity):
        year = int(m.group(1))
        # 检查上下文是否涉及年报
        start = max(0, m.start() - 50)
        end = min(len(activity), m.end() + 100)
        context = activity[start:end]
        if any(kw in context for kw in ['报告', '披露', '财务', '业绩', '利润', '收入']):
            years.add(year)

    # 模式5："XXXX年年报"
    for m in re.finditer(r'(20\d{2})年年报', activity):
        years.add(int(m.group(1)))

    # 模式6：业绩预告/快报涉及的年份
    # "XXXX年业绩预告" "XXXX年度业绩"
    for m in re.finditer(r'(20\d{2})年(?:度)?(?:业绩|净利润|营业收入)', activity):
        year = int(m.group(1))
        # 检查是否涉及年报
        start = max(0, m.start() - 30)
        end = min(len(activity), m.end() + 100)
        context = activity[start:end]
        if any(kw in context for kw in ['年度报告', '年报', '业绩预告', '业绩快报', '修正']):
            years.add(year)

    # 模式7：虚增/虚减 + 年份
    for m in re.finditer(r'(20\d{2})年(?:度)?(?:.*?(?:虚增|虚减|少计|多计|少确认|多确认))', activity):
        years.add(int(m.group(1)))

    # 模式8："XXXX年年报/年度报告" + 虚假记载
    for m in re.finditer(r'(20\d{2})年', activity):
        year = int(m.group(1))
        start = max(0, m.start() - 20)
        end = min(len(activity), m.end() + 150)
        context = activity[start:end]
        if any(kw in context for kw in ['虚假记载', '虚增', '虚减', '少计', '多计', '不准确', '不真实']):
            # 检查是否是报告年份
            if any(kw in context for kw in ['年报', '年度报告', '报告', '披露']):
                years.add(year)

    # 模式9：范围 "XXXX年至XXXX年"
    for m in re.finditer(r'(20\d{2})年至(20\d{2})年', activity):
        start_year = int(m.group(1))
        end_year = int(m.group(2))
        # 检查上下文是否涉及年报
        start = max(0, m.start() - 50)
        end = min(len(activity), m.end() + 100)
        context = activity[start:end]
        if any(kw in context for kw in ['报告', '披露', '财务', '虚假', '虚增']):
            for y in range(start_year, end_year + 1):
                years.add(y)

    # 模式10："XXXX-XXXX年"
    for m in re.finditer(r'(20\d{2})-(20\d{2})年', activity):
        start_year = int(m.group(1))
        end_year = int(m.group(2))
        start = max(0, m.start() - 50)
        end_pos = min(len(activity), m.end() + 100)
        context = activity[start:end_pos]
        if any(kw in context for kw in ['报告', '披露', '财务', '虚假', '虚增']):
            for y in range(start_year, end_year + 1):
                years.add(y)

    # 过滤
    years = {y for y in years if 2000 <= y <= 2026}

    return sorted(list(years)) if years else None


# ============================================================
# 年报相关性判断
# ============================================================

def is_annual_report_related(activity, violation_type):
    """判断违规行为是否导致年度报告中的信息存在错误或遗漏。"""
    if not activity:
        return False, "无Activity文本"

    # ---- 排除规则 ----

    # 1. 内幕交易
    if violation_type and '内幕交易' in violation_type:
        if not any(kw in violation_type for kw in ['虚假记载', '虚构利润', '欺诈上市', '重大遗漏']):
            # 检查文本是否有明确的年报问题
            if not re.search(r'(?:年度报告|年报).*(?:虚假|虚增|虚减|不准确|不真实)', activity):
                return False, "内幕交易，不影响年报"

    # 2. 违规买卖股票
    if violation_type and '违规买卖股票' in violation_type:
        if not any(kw in violation_type for kw in ['虚假记载', '虚构利润', '欺诈上市']):
            if not re.search(r'(?:年度报告|年报).*(?:虚假|虚增|虚减|不准确)', activity):
                return False, "违规买卖股票，不影响年报"

    # 3. 仅涉及半年报/季报
    if ('半年度报告' in activity or '半年报' in activity or '季度报告' in activity or '季报' in activity):
        if '年度报告' not in activity and '年报' not in activity:
            return False, "仅涉及半年报/季报"

    # 4. 未缴或少缴税款
    if violation_type and '未缴或少缴税款' in violation_type:
        if '年度报告' not in activity and '年报' not in activity:
            return False, "未缴税款，未涉及年报"

    # 5. 安全生产事故
    if '安全事故' in activity and '安全生产' in activity:
        if '年度报告' not in activity and '年报' not in activity:
            return False, "安全生产事故"

    # 6. 化妆品/药品/产品质量
    if any(kw in activity for kw in ['化妆品', '药品销售', '生产工艺', '配方投料']):
        if '年度报告' not in activity and '年报' not in activity:
            if not any(kw in violation_type for kw in ['虚假记载', '虚构利润', '重大遗漏']):
                return False, "产品质量违规"

    # 7. 董事未出席会议
    if '未出席' in activity and '董事会' in activity and '年度报告' not in activity and '年报' not in activity:
        return False, "董事未出席"

    # 8. 增持承诺未履行
    if '增持' in activity and '承诺' in activity and '年度报告' not in activity and '年报' not in activity:
        if not any(kw in violation_type for kw in ['虚假记载', '虚构利润']):
            return False, "增持承诺未履行"

    # 9. 工程违法转包
    if '违法转包' in activity:
        return False, "工程违法转包"

    # 10. 垄断/反垄断
    if '垄断' in activity or '滥用市场支配地位' in activity:
        return False, "垄断行为"

    # 11. 安全生产/人员资质
    if '安全教育' in activity and '技术培训' in activity:
        return False, "安全生产培训"

    # 12. 期货/融资融券业务
    if '融资融券业务' in activity or '期货营业部' in activity:
        return False, "期货/融资融券业务"

    # ---- 纳入规则 ----

    # 规则1：违规类型直接暗示年报问题
    annual_report_violation_types = ['虚假记载', '虚构利润', '欺诈上市', '一般会计处理不当']
    for kw in annual_report_violation_types:
        if violation_type and kw in violation_type:
            # 确认涉及年报
            if re.search(r'(?:年度报告|年报|年度|财务报表|业绩)', activity):
                return True, f"违规类型'{kw}'涉及年报"

    # 规则2：文本中明确提到年度报告有问题
    if re.search(r'(?:年度报告|年报).*(?:虚假记载|虚增|虚减|少计|多计|不准确|不真实|重大遗漏|误导性|不完整|错误|未披露|隐瞒)', activity):
        return True, "年度报告存在虚假/遗漏"

    # 规则3：业绩预告/快报与年报差异
    if re.search(r'(?:业绩预告|业绩快报).*(?:年度报告|年报)', activity):
        if re.search(r'(?:差异|修正|更正|不准确|不及时)', activity):
            return True, "业绩预告/快报与年报差异"

    if re.search(r'(?:年度报告|年报).*(?:业绩预告|业绩快报)', activity):
        if re.search(r'(?:差异|修正|更正|不准确)', activity):
            return True, "业绩预告/快报与年报差异"

    # 规则4：招股说明书虚假记载
    if '招股说明书' in activity and re.search(r'(?:虚假记载|虚增|虚减|编造|隐瞒)', activity):
        return True, "招股说明书虚假记载"

    # 规则5：涉及年度财务数据错误
    if re.search(r'(?:年度报告|年报).*(?:虚增|虚减|少计|多计)', activity):
        return True, "年度财务数据错误"

    # 规则6：年报中关联交易/关联方披露问题
    if re.search(r'(?:年度报告|年报).*(?:关联交易|关联方|关联关系).*(?:未披露|不准确|不完整|重大遗漏)', activity):
        return True, "年报关联交易披露问题"

    # 规则7：募集资金问题涉及年报
    if re.search(r'募集资金.*(?:年度报告|年报).*(?:未披露|不准确|不完整)', activity):
        return True, "年报募集资金披露问题"

    # 规则8：定期报告营业收入/利润不准确
    if re.search(r'(?:年度报告|年报).*(?:营业收入|净利润|利润).*(?:不准确|不真实|虚假)', activity):
        return True, "年报财务数据不准确"

    # 规则9：未按规定披露 + 年报
    if re.search(r'(?:年度报告|年报).*(?:未.*披露|未按规定)', activity):
        return True, "年报未按规定披露"

    # 规则10：年报 + 违规担保/资金占用
    if re.search(r'(?:年度报告|年报).*(?:违规担保|资金占用|担保)', activity):
        return True, "年报违规担保/资金占用"

    # 规则11：处置子公司/会计差错 + 年报
    if re.search(r'(?:年度报告|年报).*(?:会计差错|更正|追溯调整|处置子公司)', activity):
        return True, "年报会计差错"

    # 规则12：年报 + 公司治理/内部控制
    if re.search(r'(?:年度报告|年报).*(?:公司治理|内部控制)', activity):
        if re.search(r'(?:不规范|缺陷|问题)', activity):
            return True, "年报公司治理问题"

    return False, "未发现年报相关问题"


# ============================================================
# 财务信息判断
# ============================================================

def is_financial_info(activity, violation_type):
    """判断年报中受影响的是财务信息还是非财务信息。"""
    if not activity:
        return False, "无Activity文本"

    # 财务相关的违规类型
    financial_violation_types = ['虚假记载', '虚构利润', '一般会计处理不当', '欺诈上市', '披露不实']
    for fvt in financial_violation_types:
        if violation_type and fvt in violation_type:
            if re.search(r'(?:营业收入|净利润|利润|收入|成本|费用|资产|负债|权益|应收账款|存货|虚增|虚减|少计|多计|少确认|多确认|提前确认|坏账|减值|折旧|财务报表|财务报告|财务数据|业绩预告|业绩快报|年报)', activity):
                return True, f"违规类型'{fvt}'涉及财务数据"

    # 关键词匹配
    financial_patterns = [
        r'虚增.*(?:营业收入|收入|利润|净利润|资产|应收账款|存货)',
        r'虚减.*(?:利润|净利润|收入)',
        r'少计提.*(?:坏账|减值|折旧)',
        r'多计提',
        r'少确认.*(?:收入|费用)',
        r'多确认.*(?:收入|费用)',
        r'提前确认.*(?:收入|利润)',
        r'推迟确认',
        r'虚假记载.*(?:利润|收入|资产)',
        r'虚构.*(?:交易|业务|收入|利润)',
        r'伪造.*(?:合同|发票|凭证|银行)',
        r'(?:营业收入|净利润|利润).*(?:不准确|不真实|虚假)',
        r'(?:收入确认|成本核算|费用冲减).*(?:不规范|不准确|不当)',
        r'坏账准备.*(?:少提|多提|不当)',
        r'资产减值.*(?:少计|多计)',
        r'商誉减值',
        r'(?:折旧|摊销).*(?:少计|多计)',
        r'会计.*(?:差错|处理).*(?:更正|不准确|不当)',
        r'收入确认.*(?:跨期|提前)',
    ]

    for pattern in financial_patterns:
        if re.search(pattern, activity):
            return True, f"涉及财务关键词: {pattern}"

    return False, "未发现明确的财务信息影响"


# ============================================================
# 会计要素识别
# ============================================================

def identify_financial_elements(activity):
    """识别受影响的会计要素（六大要素）。"""
    if not activity:
        return None, []

    element_mapping = {
        '资产': [
            '应收账款', '应收票据', '其他应收款', '预付账款',
            '存货', '固定资产', '无形资产', '在建工程', '商誉',
            '长期投资', '股权投资', '投资性房地产', '开发支出',
            '货币资金', '银行存款', '资金', '资金占用',
            '坏账', '减值准备', '资产减值',
            '委托理财', '担保', '质押',
            '往来', '拆借', '保证金',
        ],
        '负债': [
            '应付账款', '应付票据', '其他应付款', '预收账款',
            '短期借款', '长期借款', '应付债券',
            '预计负债', '递延收益',
            '担保', '债务', '贷款',
        ],
        '所有者权益': [
            '实收资本', '资本公积', '盈余公积', '未分配利润',
            '所有者权益', '净资产', '股东权益',
            '配股', '送股', '转增',
        ],
        '收入': [
            '营业收入', '主营业务收入', '其他业务收入',
            '销售收入', '确认收入',
            '投资收益', '利息收入',
            '提前确认', '推迟确认',
        ],
        '费用': [
            '营业成本', '主营业务成本', '管理费用', '销售费用',
            '财务费用', '研发费用', '所得税费用',
            '成本', '折旧', '摊销',
            '坏账准备', '减值准备', '减值损失',
            '信用减值', '资产减值',
            '税金',
        ],
        '利润': [
            '净利润', '利润总额', '营业利润',
            '归属于上市公司股东的净利润',
            '扣非净利润', '扣除非经常性损益后的净利润',
            '综合收益',
        ],
    }

    found_elements = set()
    year_elements = {}

    # 按年份提取要素
    year_pattern = r'(20\d{2})年'
    year_matches = list(re.finditer(year_pattern, activity))

    for match in year_matches:
        year = int(match.group(1))
        if year < 2000 or year > 2026:
            continue

        start_pos = match.start()
        next_year_match = re.search(r'(20\d{2})年', activity[match.end():])
        if next_year_match:
            end_pos = match.end() + next_year_match.start()
        else:
            end_pos = len(activity)

        segment = activity[start_pos:end_pos]

        for element, keywords in element_mapping.items():
            for kw in keywords:
                if kw in segment:
                    if year not in year_elements:
                        year_elements[year] = set()
                    year_elements[year].add(element)
                    found_elements.add(element)
                    break

    # 全局搜索
    if not found_elements:
        for element, keywords in element_mapping.items():
            for kw in keywords:
                if kw in activity:
                    found_elements.add(element)
                    break

    # 格式化输出
    if year_elements:
        ann_fin_info = []
        for year in sorted(year_elements.keys()):
            elements = sorted(list(year_elements[year]))
            ann_fin_info.append({"year": year, "elements": elements})
        return ann_fin_info, list(found_elements)
    elif found_elements:
        return None, list(found_elements)
    else:
        return None, []


# ============================================================
# 第三方配合造假识别
# ============================================================

def identify_third_party(activity, violation_type):
    """识别是否存在第三方配合财务舞弊。"""
    if not activity:
        return False, None

    # 排除：审计机构的正常审计工作（非配合造假）
    if re.search(r'(?:会计师事务所|审计机构).*(?:审计|执业|核查)', activity):
        if not re.search(r'(?:配合|串通|合谋|虚假.*审计|出具虚假)', activity):
            # 审计机构的正常工作不算第三方配合
            pass

    # 第三方配合的关键词
    collusion_patterns = [
        r'配合签订虚假',
        r'配合开具.*(?:虚假|不实)',
        r'签订虚假.*(?:采购|销售|购销).*合同',
        r'虚开.*(?:发票|增值税|销售发票)',
        r'虚假.*(?:验收|证明)',
        r'伪造.*(?:凭证|合同|发票)',
        r'资金.*(?:回流|循环)',
        r'出借账户',
        r'代持股份',
        r'配合.*(?:造假|舞弊|虚增)',
        r'协助.*(?:造假|舞弊)',
        r'配合虚构',
        r'串通.*(?:造假|舞弊|虚增)',
        r'合谋.*(?:造假|舞弊)',
        r'共谋',
    ]

    has_collusion = False
    for pattern in collusion_patterns:
        if re.search(pattern, activity):
            has_collusion = True
            break

    if not has_collusion:
        return False, None

    # 提取第三方名称
    third_parties = []

    # 更精确的名称提取
    # 模式：与XXX签订虚假合同
    name_patterns = [
        r'(?:与|向)([一-龥]{2,25}(?:有限公司|股份有限公司|集团)).*?(?:签订|签署).*?(?:虚假|不实)',
        r'([一-龥]{2,25}(?:有限公司|股份有限公司|集团)).*?(?:配合|协助).*?(?:签订|开具|提供).*?(?:虚假|不实)',
        r'供应商([一-龥]{2,25}(?:有限公司|股份有限公司|集团))',
        r'客户([一-龥]{2,25}(?:有限公司|股份有限公司|集团))',
        r'([一-龥]{2,25}(?:有限公司|股份有限公司)).*?(?:提供|出具).*?(?:虚假|不实)',
        r'自然人([一-龥]{2,6}).*?(?:代持|配合)',
    ]

    for pattern in name_patterns:
        matches = re.findall(pattern, activity)
        for match in matches:
            name = match.strip()
            if name and len(name) >= 2:
                # 排除上市公司自身
                if any(kw in name for kw in ['本公司', '公司自身', '上市公司']):
                    continue

                # 判断类型
                ptype = "其他企业"
                ctx = activity[max(0, activity.find(name)-30):activity.find(name)+len(name)+30]
                if '供应商' in ctx:
                    ptype = "供应商"
                elif '客户' in ctx:
                    ptype = "客户"
                elif '银行' in ctx:
                    ptype = "银行/金融机构"
                elif '会计师' in name:
                    ptype = "会计师事务所"
                elif '评估' in name:
                    ptype = "评估机构"
                elif '券商' in name or '证券' in name:
                    ptype = "券商/保荐机构"
                elif '自然人' in ctx:
                    ptype = "自然人"

                # 获取角色描述
                role = "配合财务舞弊"
                if '虚假合同' in activity[activity.find(name):activity.find(name)+100]:
                    role = "签订虚假合同"
                elif '虚假验收' in activity[activity.find(name):activity.find(name)+100]:
                    role = "配合开具虚假验收证明"
                elif '资金回流' in activity[activity.find(name):activity.find(name)+100]:
                    role = "配合资金回流"
                elif '虚开' in activity[activity.find(name):activity.find(name)+100]:
                    role = "虚开票据"
                elif '代持' in activity[activity.find(name):activity.find(name)+100]:
                    role = "代持股份"

                if name not in [p.get('name') for p in third_parties]:
                    third_parties.append({
                        "name": name,
                        "type": ptype,
                        "role": role
                    })

    if third_parties:
        return True, third_parties

    return False, None


# ============================================================
# 主处理函数
# ============================================================

def process_row(activity, violation_type):
    """处理单行数据。"""
    ann_related = 0
    ann_year = None
    ann_fin_flag = None
    ann_fin_info = None
    third_party_flag = None
    third_party_list = None

    # Step 1: 判断年报相关性
    is_related, reason = is_annual_report_related(activity, violation_type)
    ann_related = 1 if is_related else 0

    if ann_related == 0:
        return ann_related, ann_year, ann_fin_flag, ann_fin_info, third_party_flag, third_party_list

    # Step 2: 提取受影响年份
    ann_year = extract_years(activity)

    # Step 3: 判断是否影响财务信息
    is_financial, fin_reason = is_financial_info(activity, violation_type)
    ann_fin_flag = 1 if is_financial else 0

    if ann_fin_flag == 0:
        return ann_related, ann_year, ann_fin_flag, ann_fin_info, third_party_flag, third_party_list

    # Step 4: 识别会计要素
    ann_fin_info, elements = identify_financial_elements(activity)

    if ann_fin_info is None and elements and ann_year:
        ann_fin_info = []
        for year in ann_year:
            ann_fin_info.append({"year": year, "elements": elements})

    # Step 5: 判断第三方配合
    has_third_party, third_party_list = identify_third_party(activity, violation_type)
    third_party_flag = 1 if has_third_party else 0

    return ann_related, ann_year, ann_fin_flag, ann_fin_info, third_party_flag, third_party_list


# ============================================================
# 主程序
# ============================================================

def main():
    input_file = 'blank_G02.xlsx'
    output_file = 'Mimo/mimo_ver.xlsx'

    print("正在读取数据...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    activity_col = None
    violation_type_col = None
    for i, h in enumerate(headers):
        if h == 'Activity':
            activity_col = i + 1
        if h == 'ViolationType':
            violation_type_col = i + 1

    print(f"Activity列: 第{activity_col}列, ViolationType列: 第{violation_type_col}列")

    results = []
    total_rows = ws.max_row - 1

    for row_idx in range(2, ws.max_row + 1):
        activity = ws.cell(row=row_idx, column=activity_col).value or ""
        violation_type = ws.cell(row=row_idx, column=violation_type_col).value or ""

        ann_related, ann_year, ann_fin_flag, ann_fin_info, third_party_flag, third_party_list = process_row(activity, violation_type)

        ann_year_str = json.dumps(ann_year, ensure_ascii=False) if ann_year else "null"
        ann_fin_info_str = json.dumps(ann_fin_info, ensure_ascii=False) if ann_fin_info else "null"
        third_party_list_str = json.dumps(third_party_list, ensure_ascii=False) if third_party_list else "null"

        results.append({
            'row': row_idx,
            'ann_related': ann_related,
            'ann_year': ann_year_str,
            'ann_fin_flag': ann_fin_flag,
            'ann_fin_info': ann_fin_info_str,
            'third_party_flag': third_party_flag,
            'third_party_list': third_party_list_str,
        })

        if (row_idx - 1) % 30 == 0 or row_idx == ws.max_row:
            print(f"  已处理 {row_idx - 1}/{total_rows} 行...")

    # 统计
    print("\n=== 标注统计 ===")
    ann_related_count = sum(1 for r in results if r['ann_related'] == 1)
    ann_fin_count = sum(1 for r in results if r['ann_fin_flag'] == 1)
    third_party_count = sum(1 for r in results if r['third_party_flag'] == 1)
    print(f"总行数: {total_rows}")
    print(f"ann_related=1: {ann_related_count}")
    print(f"ann_fin_flag=1: {ann_fin_count}")
    print(f"third_party_flag=1: {third_party_count}")

    # 输出Excel
    print(f"\n正在生成: {output_file}")
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'sheet1'

    out_headers = ['Activity', 'ann_related', 'ann_year', 'ann_fin_flag', 'ann_fin_info', 'third_party_flag', 'third_party_list']
    for col, header in enumerate(out_headers, 1):
        out_ws.cell(row=1, column=col, value=header)

    for row_idx, result in enumerate(results, 2):
        activity = ws.cell(row=row_idx, column=activity_col).value or ""
        out_ws.cell(row=row_idx, column=1, value=activity)
        out_ws.cell(row=row_idx, column=2, value=result['ann_related'])
        out_ws.cell(row=row_idx, column=3, value=result['ann_year'])
        out_ws.cell(row=row_idx, column=4, value=result['ann_fin_flag'])
        out_ws.cell(row=row_idx, column=5, value=result['ann_fin_info'])
        out_ws.cell(row=row_idx, column=6, value=result['third_party_flag'])
        out_ws.cell(row=row_idx, column=7, value=result['third_party_list'])

    out_ws.column_dimensions['A'].width = 80
    out_ws.column_dimensions['B'].width = 12
    out_ws.column_dimensions['C'].width = 18
    out_ws.column_dimensions['D'].width = 14
    out_ws.column_dimensions['E'].width = 60
    out_ws.column_dimensions['F'].width = 16
    out_ws.column_dimensions['G'].width = 60

    out_wb.save(output_file)
    print(f"输出文件已保存: {output_file}")

    # 输出ann_related=1的详细结果
    print("\n=== ann_related=1 的行 ===")
    for result in results:
        if result['ann_related'] == 1:
            print(f"Row {result['row']}: ann_year={result['ann_year']}, "
                  f"ann_fin_flag={result['ann_fin_flag']}, "
                  f"third_party_flag={result['third_party_flag']}")


if __name__ == '__main__':
    main()
